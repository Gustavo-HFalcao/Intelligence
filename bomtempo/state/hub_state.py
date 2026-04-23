"""
HubState — State for Hub de Operações sub-modules:
  • Cronograma: CRUD de atividades, edição inline, gestão de dependências
  • Auditoria: Bolsões de imagens (Equipe, Falhas, Ferramentas, Gerais) com lightbox
  • Timeline: Log de eventos/registros do projeto, comentários, @mentions

Supabase tables required (see schema below):
  - hub_atividades: id, contrato, fase_macro, fase, atividade, responsavel,
                    inicio_previsto, termino_previsto, conclusao_pct, critico,
                    dependencia, observacoes, created_at, updated_at, created_by
  - hub_auditoria_imgs: id, contrato, categoria, url, legenda, data_captura, autor, created_at
  - hub_timeline: id, contrato, tipo, titulo, descricao, autor, created_at, mencoes (jsonb)
"""
import logging
from typing import Any, Dict, List
import reflex as rx

from bomtempo.core.supabase_client import sb_select, sb_insert, sb_update, sb_delete
from bomtempo.core.audit_logger import audit_log, AuditCategory
from bomtempo.core.executors import (
    get_ai_executor,
    get_db_executor,
    get_http_executor,
)

logger = logging.getLogger(__name__)

from datetime import timezone, timedelta
_BRT = timezone(timedelta(hours=-3))


def _utc_to_brt(ts: str) -> str:
    """Convert ISO UTC timestamp → BRT (UTC-3), formatted as DD/MM/YYYY HH:MM."""
    if not ts:
        return ""
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(ts.replace("Z", "+00:00")[:32])
        brt = dt.astimezone(_BRT)
        return brt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ts[:16].replace("T", " ")


def _utc_date_to_br(ts: str) -> str:
    """Convert ISO date string YYYY-MM-DD → DD/MM/YYYY."""
    if not ts:
        return ""
    try:
        parts = ts[:10].split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return ts[:10]

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

ENTRY_TYPES = ["Atualização", "Marco", "Reunião", "Decisão", "Alerta", "Falha", "Documento", "Custo"]

AUDIT_CATEGORIES = [
    {"slug": "equipe",       "label": "Equipe com EPI",    "icon": "hard-hat",     "color": "#22c55e"},
    {"slug": "falhas",       "label": "Falhas & Logs",     "icon": "alert-triangle","color": "#EF4444"},
    {"slug": "ferramentas",  "label": "Ferramentas",       "icon": "wrench",       "color": "#2A9D8F"},
    {"slug": "gerais",       "label": "Imagens Gerais",    "icon": "image",        "color": "#C98B2A"},
]

FASE_COLORS: Dict[str, str] = {
    "civil":       "#C98B2A",
    "elétrica":    "#3B82F6",
    "eletrica":    "#3B82F6",
    "hidráulica":  "#2A9D8F",
    "hidraulica":  "#2A9D8F",
    "estrutural":  "#E89845",
    "mecânica":    "#A855F7",
    "mecanica":    "#A855F7",
    "licenciamento": "#64748B",
    "aprovações":  "#64748B",
    "aprovacoes":  "#64748B",
}


def _fase_color(fase: str) -> str:
    return FASE_COLORS.get(fase.lower().strip(), "#889999")


_DIAS_MAP = {"seg": 0, "ter": 1, "qua": 2, "qui": 3, "sex": 4, "sab": 5, "sáb": 5, "dom": 6}


def _parse_dias_uteis(dias_str: str) -> set:
    """Convert 'seg,ter,qua,qui,sex' to a set of weekday ints (0=Mon…6=Sun).
    Falls back to Mon-Fri if the string is empty or unrecognized."""
    if not dias_str:
        return {0, 1, 2, 3, 4}
    result = set()
    for d in dias_str.split(","):
        key = d.strip().lower()
        if key in _DIAS_MAP:
            result.add(_DIAS_MAP[key])
    return result if result else {0, 1, 2, 3, 4}


def _add_working_days(start_iso: str, days: int, working_days: set = None) -> str:
    """Return ISO date string for an activity that starts on start_iso and lasts `days` working days.

    The start day itself counts as day 1. So dias=1 → termino=inicio; dias=2 → termino=next working day.
    working_days: set of weekday ints (0=Mon…6=Sun). Defaults to Mon-Fri.
    """
    from datetime import date, timedelta
    if working_days is None:
        working_days = {0, 1, 2, 3, 4}
    try:
        current = date.fromisoformat(start_iso[:10])
        if not working_days or days <= 0:
            return start_iso
        # Day 1 is the start date itself; advance (days-1) additional working days
        added = 1
        while added < days:
            current += timedelta(days=1)
            if current.weekday() in working_days:
                added += 1
        return current.isoformat()
    except Exception:
        return start_iso


def _count_working_days(start_iso: str, end_iso: str, working_days: set = None) -> int:
    """Count working days between start and end dates (inclusive).
    Day 1 is the start date itself, matching _add_working_days convention."""
    from datetime import date, timedelta
    if working_days is None:
        working_days = {0, 1, 2, 3, 4}
    try:
        d0 = date.fromisoformat(start_iso[:10])
        d1 = date.fromisoformat(end_iso[:10])
        if d1 < d0:
            return 0
        count = 0
        cur = d0
        while cur <= d1:
            if cur.weekday() in working_days:
                count += 1
            cur += timedelta(days=1)
        return count
    except Exception:
        return 0


def _norm_str(v: object, fallback: str = "") -> str:
    if v is None or str(v) in ("None", "NaT", "nan", ""):
        return fallback
    return str(v)


def _recalc_macro_dates(macro_id: str, contrato: str, client_id: str) -> None:
    """After saving a micro, update its parent macro's inicio/termino from children dates."""
    try:
        children = sb_select("hub_atividades", filters={"parent_id": macro_id, "contrato": contrato}, limit=200)
        starts = [r["inicio_previsto"] for r in children if r.get("inicio_previsto")]
        ends = [r["termino_previsto"] for r in children if r.get("termino_previsto")]
        if starts and ends:
            sb_update("hub_atividades", filters={"id": macro_id}, data={
                "inicio_previsto": min(starts),
                "termino_previsto": max(ends),
            })
    except Exception as e:
        logger.warning(f"_recalc_macro_dates error: {e}")


def _log_schedule_diff_async(
    contrato: str,
    atividade_id: str,
    atividade_nome: str,
    old_row: dict,
    new_row: dict,
    autor: str,
    client_id: str,
) -> None:
    """Fire-and-forget: record a full diff of changed fields into hub_cronograma_log
    + summary entry in hub_timeline with AI impact note (best-effort).

    Tracked fields:
      inicio_previsto, termino_previsto, conclusao_pct, responsavel,
      peso_pct, critico, nivel, fase_macro, fase, observacoes,
      total_qty, unidade, dias_planejados
    """
    import threading

    FIELD_LABELS = {
        "inicio_previsto":  "Início",
        "termino_previsto": "Término",
        "conclusao_pct":    "Conclusão %",
        "responsavel":      "Responsável",
        "peso_pct":         "Peso %",
        "critico":          "Crítico",
        "nivel":            "Nível",
        "fase_macro":       "Fase Macro",
        "fase":             "Fase",
        "observacoes":      "Observações",
        "total_qty":        "Qtd Total",
        "unidade":          "Unidade",
        "dias_planejados":  "Dias Planejados",
        "status_atividade": "Status",
        "tipo_medicao":     "Tipo Medição",
    }

    def _fmt(field: str, v) -> str:
        v = str(v or "")
        if field in ("inicio_previsto", "termino_previsto", "data_inicio_real", "data_fim_real", "data_fim_prevista"):
            return _utc_date_to_br(v) or "—"
        if field == "critico":
            return "Sim" if v.upper() in ("TRUE", "1", "SIM", "YES") else "Não"
        return v or "—"

    def _run():
        try:
            diffs = []
            for field in FIELD_LABELS:
                old_v = str(old_row.get(field, "") or "")
                new_v = str(new_row.get(field, "") or "")
                if old_v != new_v:
                    diffs.append((field, old_v, new_v))

            if not diffs:
                return

            # Bulk insert all diff rows in one HTTP request
            log_rows = [
                {
                    "contrato":       contrato,
                    "atividade_id":   atividade_id or None,
                    "atividade_nome": atividade_nome[:120],
                    "campo":          field,
                    "valor_anterior": _fmt(field, old_v)[:500],
                    "valor_novo":     _fmt(field, new_v)[:500],
                    "autor":          autor or "sistema",
                    "client_id":      client_id or None,
                }
                for field, old_v, new_v in diffs
            ]
            try:
                from bomtempo.core.supabase_client import sb_bulk_upsert as _sb_bulk
                _sb_bulk("hub_cronograma_log", log_rows, on_conflict="id")
            except Exception:
                pass

            # Human-readable summary for timeline
            date_fields = {"inicio_previsto", "termino_previsto"}
            date_diffs = [(f, o, n) for f, o, n in diffs if f in date_fields]
            other_diffs = [(f, o, n) for f, o, n in diffs if f not in date_fields]

            parts = []
            for f, o, n in date_diffs:
                parts.append(f"{FIELD_LABELS[f]}: {_fmt(f, o)} → {_fmt(f, n)}")
            for f, o, n in other_diffs[:4]:  # cap at 4 non-date changes per entry
                label = FIELD_LABELS.get(f, f)
                parts.append(f"{label}: {_fmt(f, o)} → {_fmt(f, n)}")

            change_summary = " | ".join(parts)
            titulo = f"[Cronograma] {atividade_nome[:60]} — {len(diffs)} campo(s) alterado(s)"

            tl_id = None
            try:
                result = sb_insert("hub_timeline", {
                    "contrato":    contrato,
                    "tipo":        "Atualização",
                    "titulo":      titulo,
                    "descricao":   change_summary,
                    "autor":       autor or "sistema",
                    "mencoes":     [],
                    "is_document": False,
                    "is_cost":     False,
                    "client_id":   client_id or None,
                })
                # sb_insert returns a dict (single row) or None
                if result and isinstance(result, dict) and result.get("id"):
                    tl_id = result["id"]
                elif result and isinstance(result, list) and result[0].get("id"):
                    tl_id = result[0]["id"]
            except Exception:
                pass

            # AI impact analysis (best-effort, enriches the timeline entry)
            try:
                from bomtempo.core.ai_client import ai_client
                diff_text = "\n".join(
                    f"- {FIELD_LABELS.get(f, f)}: '{_fmt(f, o)}' → '{_fmt(f, n)}'"
                    for f, o, n in diffs
                )
                msg = [{"role": "user", "content": (
                    f"Atividade de obra: '{atividade_nome}'\n"
                    f"Alterações registradas:\n{diff_text}\n\n"
                    f"Em 1-2 frases objetivas: qual o impacto dessas mudanças no cronograma? "
                    f"Considere dependências, prazo contratual e caminho crítico."
                )}]
                ai_note = ai_client.query(msg)
                if ai_note:
                    # Update timeline entry with AI note
                    if tl_id:
                        sb_update("hub_timeline", filters={"id": tl_id},
                                  data={"descricao": f"{change_summary}\n\n[agente] {ai_note}"})
                    # Also store impacto in the log rows for this batch
                    try:
                        from datetime import datetime, timezone
                        # Update the most recent log row for this atividade
                        sb_update(
                            "hub_cronograma_log",
                            filters={"atividade_id": atividade_id, "autor": autor},
                            data={"ai_impacto": ai_note[:1000]},
                        )
                    except Exception:
                        pass
            except Exception:
                pass

        except Exception as ex:
            logger.warning(f"_log_schedule_diff_async error: {ex}")

    threading.Thread(target=_run, daemon=True).start()


# Keep legacy name as thin wrapper for backward compat
def _log_schedule_change_async(
    contrato: str, atividade: str,
    old_inicio: str, new_inicio: str,
    old_termino: str, new_termino: str,
    autor: str, client_id: str,
) -> None:
    """Legacy wrapper — delegates to _log_schedule_diff_async."""
    _log_schedule_diff_async(
        contrato=contrato,
        atividade_id="",
        atividade_nome=atividade,
        old_row={"inicio_previsto": old_inicio, "termino_previsto": old_termino},
        new_row={"inicio_previsto": new_inicio, "termino_previsto": new_termino},
        autor=autor,
        client_id=client_id,
    )


def _norm_pct(v: object) -> str:
    try:
        return str(int(float(v or 0)))
    except (ValueError, TypeError):
        return "0"


# ─────────────────────────────────────────────────────────────────────────────
# State
# ─────────────────────────────────────────────────────────────────────────────

def _working_days_between(d_start, d_end) -> int:
    """Conta dias úteis (seg-sex) entre duas datas, inclusive d_start, exclusive d_end."""
    from datetime import timedelta
    if d_end <= d_start:
        return 0
    total = (d_end - d_start).days
    # Semanas completas
    weeks, rem = divmod(total, 7)
    wd = weeks * 5
    # Dias restantes: conta seg-sex começando do weekday de d_start
    start_wd = d_start.weekday()  # 0=seg, 6=dom
    for i in range(rem):
        if (start_wd + i) % 7 < 5:
            wd += 1
    return max(0, wd)


def _add_working_days_simple(d_start, n: int):
    """Avança n dias úteis a partir de d_start."""
    from datetime import timedelta
    current = d_start
    added = 0
    while added < n:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current


def _compute_forecast_rows(cron_rows: list, reference_date=None) -> list:
    """
    Função pura: enriquece micro-atividades (e suas subs) com campos de forecast.
    Executada em run_in_executor (thread pool) pelo load_cronograma.

    reference_date: âncora temporal para cálculos de SPI/EAC/desvio.
      - None (padrão) → date.today() — hub ao vivo
      - date explícita → usa essa data como "hoje efetivo", ex: data do RDO sendo processado
      Atividades com last_rdo_date preenchida usam min(reference_date, last_rdo_date)
      para evitar penalizar atividades cujo RDO ainda não chegou hoje.

    Correções v2:
    - pct da micro recalculado a partir das subs (mesmo bug que filtered_cron_rows já corrigia,
      mas o forecast lia o valor bruto do banco)
    - dias úteis reais usados em todo o cálculo (não aproximação * 5/7)
    - threshold "sem_dados": só quando exec_qty == 0 (não mais dias_decorridos < 3)
    - tendência override: se EAC já passou do termino_previsto, nunca pode ser "acima"
    - EAC projetado em dias úteis, não corridos
    - pct_esperado baseado em dias úteis decorridos vs dias úteis planejados
    - subs incluídas no resultado, agrupadas abaixo da sua micro
    """
    from datetime import date
    today = reference_date if reference_date is not None else date.today()
    tol = 10.0

    # ── Pré-indexar subs por parent_id ───────────────────────────────────────
    subs_by_parent: dict = {}
    for r in cron_rows:
        if r.get("nivel", "macro") not in ("macro", "micro", ""):
            pid = r.get("parent_id", "")
            if pid:
                subs_by_parent.setdefault(pid, []).append(r)

    result = []

    for r in cron_rows:
        if r.get("nivel", "macro") != "micro":
            continue

        total_qty_s = r.get("total_qty",  "0") or "0"
        exec_qty_s  = r.get("exec_qty",   "0") or "0"
        dias_plan_s = r.get("dias_planejados", "0") or "0"
        micro_id    = r.get("id", "")

        try:
            total_qty = float(total_qty_s)
            exec_qty  = float(exec_qty_s)
            dias_plan = int(dias_plan_s)
        except (ValueError, TypeError):
            continue

        inicio_iso  = r.get("inicio_iso", "")
        termino_iso = r.get("termino_iso", "")
        unidade     = r.get("unidade", "")

        # ── FIX 3: recalcular pct da micro a partir das subs se existirem ──
        sub_list = subs_by_parent.get(micro_id, [])
        if sub_list:
            sub_peso_total = sum(int(s.get("peso_pct", "0") or "0") for s in sub_list)
            if sub_peso_total > 0:
                sub_wpct = sum(
                    int(s.get("conclusao_pct", "0") or "0") * int(s.get("peso_pct", "0") or "0")
                    for s in sub_list
                ) / sub_peso_total
            else:
                sub_wpct = sum(int(s.get("conclusao_pct", "0") or "0") for s in sub_list) / len(sub_list)
            pct = round(sub_wpct)
        else:
            pct_s = r.get("conclusao_pct", "0") or "0"
            try:
                pct = int(pct_s)
            except (ValueError, TypeError):
                pct = 0

        prod_plan = total_qty / dias_plan if total_qty > 0 and dias_plan > 0 else 0.0

        d_inicio = d_termino = None
        if inicio_iso and len(inicio_iso) >= 10:
            try: d_inicio = date.fromisoformat(inicio_iso[:10])
            except ValueError: pass
        if termino_iso and len(termino_iso) >= 10:
            try: d_termino = date.fromisoformat(termino_iso[:10])
            except ValueError: pass

        # ── FIX 1 + 6: dias úteis reais, não aproximação ──────────────────
        # Se last_rdo_date disponível: usar como teto de "hoje efetivo" por atividade
        # Garante que preenchimento retroativo não avança dias além da data do RDO
        _lrd_s = r.get("last_rdo_date", "") or ""
        if _lrd_s and len(_lrd_s) >= 10:
            try:
                _lrd = date.fromisoformat(_lrd_s[:10])
                _effective_today = min(today, _lrd)
            except ValueError:
                _effective_today = today
        else:
            _effective_today = today

        if d_inicio and d_inicio <= _effective_today:
            dias_uteis_decorridos = _working_days_between(d_inicio, _effective_today + __import__('datetime').timedelta(days=1))
            # dia_atual = dia útil corrente dentro do plano
            dia_atual = min(dias_uteis_decorridos, dias_plan)
        else:
            dias_uteis_decorridos = 0
            dia_atual = 0

        # FIX 6: pct_esperado baseado em dias úteis, não corridos
        pct_esperado = round(min(100.0, dia_atual / dias_plan * 100), 1) if dias_plan > 0 and dia_atual > 0 else 0.0
        # EAC projetado a partir de _effective_today (não date.today())
        _eac_base = _effective_today

        # prod_real: baseado em dias úteis reais decorridos (min 1 para evitar div/0)
        prod_real = exec_qty / max(1, dias_uteis_decorridos) if exec_qty > 0 else 0.0

        desvio_pct = round((prod_real - prod_plan) / prod_plan * 100, 1) if prod_plan > 0 else 0.0

        exec_label = ""
        if total_qty > 0 and unidade:
            exec_label = f"{exec_qty:.1f} de {total_qty:.1f} {unidade}"
        elif total_qty > 0:
            exec_label = f"{exec_qty:.1f} de {total_qty:.1f}"

        # ── FIX 4: EAC em dias úteis — ancorando em _effective_today ─────────
        # Usa _effective_today (= min(today, last_rdo_date)) como base do EAC:
        # preenchimento retroativo não projeta a partir de "hoje real" mas da data do RDO
        data_fim_prev_str = ""
        desvio_dias = 0
        if prod_real > 0 and total_qty > 0 and pct < 100:
            saldo = max(0.0, total_qty - exec_qty)
            dias_restantes_uteis = max(1, round(saldo / prod_real))
            fim_prev = _add_working_days_simple(_eac_base, dias_restantes_uteis)
            data_fim_prev_str = fim_prev.isoformat()
            if d_termino:
                # desvio em dias úteis também
                if fim_prev > d_termino:
                    desvio_dias = _working_days_between(d_termino, fim_prev)
                elif fim_prev < d_termino:
                    desvio_dias = -_working_days_between(fim_prev, d_termino)

        # ── FIX 1: threshold sem_dados: só quando realmente sem produção ──
        # FIX 2: override acima→abaixo se EAC já ultrapassou termino_previsto
        prazo_estourado = (
            data_fim_prev_str != "" and d_termino is not None
            and desvio_dias > 0
        )

        if exec_qty == 0:
            tendencia = "sem_dados"
        elif pct >= 100:
            tendencia = "concluida"
        elif desvio_pct >= tol and not prazo_estourado:
            tendencia = "acima"
        elif desvio_pct <= -tol or prazo_estourado:
            tendencia = "abaixo"
        else:
            tendencia = "dentro"

        # ── FIX 5: magnitude do desvio para display ───────────────────────
        if abs(desvio_pct) < 1:
            desvio_label = "no ritmo"
        elif abs(desvio_pct) < tol:
            desvio_label = f"{desvio_pct:+.0f}%"
        elif abs(desvio_pct) < 50:
            desvio_label = f"{desvio_pct:+.0f}%"
        else:
            desvio_label = f"{desvio_pct:+.0f}%"  # sempre mostrar

        result.append(dict(
            r,
            conclusao_pct=str(pct),         # pct recalculado pelas subs
            _prod_planejada=f"{prod_plan:.1f}",
            _prod_real=f"{prod_real:.1f}",
            _desvio_pct=f"{desvio_pct:+.1f}",
            _desvio_label=desvio_label,
            _tendencia=tendencia,
            _prazo_estourado="1" if prazo_estourado else "0",
            _data_fim_prevista=_utc_date_to_br(data_fim_prev_str) if data_fim_prev_str else "—",
            _desvio_dias=str(desvio_dias),
            _saldo_qty=f"{max(0.0, total_qty - exec_qty):.1f}",
            _dia_atual=str(dia_atual),
            _total_dias=str(dias_plan),
            _pct_esperado=str(pct_esperado),
            _exec_label=exec_label,
            _has_subs="1" if sub_list else "0",
            _sub_count=str(len(sub_list)),
        ))

        # ── Subs agrupadas logo abaixo da micro ───────────────────────────
        for s in sub_list:
            sub_pct_s = s.get("conclusao_pct", "0") or "0"
            try:
                sub_pct = int(sub_pct_s)
            except (ValueError, TypeError):
                sub_pct = 0
            sub_peso = s.get("peso_pct", "0") or "0"

            # Tendência da sub: simples — só baseada no pct vs pct_esperado da micro
            if sub_pct >= 100:
                sub_tend = "concluida"
            elif dia_atual > 0 and pct_esperado > 0:
                sub_desvio = sub_pct - pct_esperado
                if sub_desvio >= tol:
                    sub_tend = "acima"
                elif sub_desvio <= -tol:
                    sub_tend = "abaixo"
                else:
                    sub_tend = "dentro"
            else:
                sub_tend = "sem_dados"

            result.append(dict(
                s,
                _nivel_display="sub",
                _tendencia=sub_tend,
                _desvio_pct="0",
                _desvio_label="",
                _prod_planejada="—",
                _prod_real="—",
                _data_fim_prevista="—",
                _desvio_dias="0",
                _saldo_qty="0",
                _dia_atual=str(dia_atual),
                _total_dias=str(dias_plan),
                _pct_esperado=str(pct_esperado),
                _exec_label="",
                _has_subs="0",
                _sub_count="0",
                _prazo_estourado="0",
                _is_sub="1",
                _sub_peso=sub_peso,
            ))

    return result


class HubState(rx.State):
    """Hub de Operações — Cronograma, Auditoria, Timeline state."""

    # ── Loading flags ────────────────────────────────────────────────────────
    cron_loading: bool = False
    audit_loading: bool = False
    timeline_loading: bool = False

    # ── Forecast cache — computado em background no load_cronograma ──────────
    # Evita recalcular cron_forecast_rows (O(n) heavy) em toda mudança de state.
    _cron_forecast_cache: List[Dict[str, str]] = []

    # Filtro ativo do painel Produtividade & Forecast
    # "execucao" | "concluida" | "prevista" | "todas"
    cron_forecast_filter: str = "execucao"

    def set_cron_forecast_filter(self, v: str):
        self.cron_forecast_filter = v

    # ── Lazy load sentinels — True após primeiro carregamento para o contrato ──
    # Evita disparar load_auditoria e load_timeline no select_project:
    # só carregam quando a tab for clicada pela primeira vez.
    _audit_loaded_contrato: str = ""
    _timeline_loaded_contrato: str = ""

    # ══════════════════════════════════════════════════════════════════════════
    # CRONOGRAMA
    # ══════════════════════════════════════════════════════════════════════════

    # List of normalized activity dicts for the selected contract
    # Keys: id, contrato, fase_macro, fase, atividade, responsavel,
    #       inicio_previsto, termino_previsto, conclusao_pct, critico,
    #       dependencia, observacoes, color
    cron_rows: List[Dict[str, str]] = []

    # Working days config for the current project (loaded from contratos.dias_uteis_semana)
    cron_working_days_str: str = "seg,ter,qua,qui,sex"
    # Current contract code (set by load_cronograma so upload handlers can read it)
    cron_contrato: str = ""

    # Filter
    cron_fase_filter: str = ""
    cron_search: str = ""
    cron_search_input: str = ""  # UI-only: updated on_change, committed on_blur/Enter
    cron_show_only_critical: bool = False

    # KPI detail popup (programado hoje / atrasadas / em risco / adiantadas)
    cron_kpi_popup: str = ""   # "" | "programado" | "atrasadas" | "em_risco" | "adiantadas" | "realizado"

    def set_cron_kpi_popup(self, v: str):
        self.cron_kpi_popup = v

    # Inline edit dialog
    cron_show_dialog: bool = False
    cron_edit_id: str = ""          # empty = new
    cron_pending_review_id: str = ""  # set when opening dialog for pending approval
    cron_edit_atividade: str = ""
    cron_edit_fase_macro: str = ""
    cron_edit_fase: str = ""
    cron_edit_responsavel: str = ""
    cron_edit_inicio: str = ""
    cron_edit_termino: str = ""
    cron_edit_pct: str = "0"
    cron_edit_critico: bool = False
    cron_edit_dependencia: str = ""      # legacy text name (kept for compat)
    cron_edit_dependencia_id: str = ""   # UUID of dependency activity
    # dep_tipo: 'sem_dep' | 'tradicional' | 'progresso'
    cron_edit_dep_tipo: str = "sem_dep"
    cron_edit_observacoes: str = ""
    cron_saving: bool = False
    cron_error: str = ""
    # Quantity tracking (#17)
    cron_edit_total_qty: str = ""    # total planned quantity (e.g. "1456")
    cron_edit_unidade: str = ""      # unit (e.g. "perfurações", "m²", "un")
    cron_edit_dias_planejados: str = ""  # working days → auto-fill termino
    # Forecast / status fields
    cron_edit_status_atividade: str = "nao_iniciada"
    cron_edit_tipo_medicao: str = "quantidade"
    # Efetivo alocado planejado para esta atividade
    cron_edit_efetivo_alocado: str = ""

    # Delete confirm
    cron_delete_id: str = ""
    cron_delete_name: str = ""
    cron_show_delete: bool = False

    # IA import from Excel/PDF
    cron_import_loading: bool = False
    cron_import_error: str = ""
    cron_import_preview: List[Dict[str, str]] = []  # proposed activities to review
    cron_import_show: bool = False
    cron_import_selected: List[str] = []  # ids of proposals to confirm
    cron_import_confidence_label: str = ""  # e.g. "Alta (92%)"

    # ══════════════════════════════════════════════════════════════════════════
    # AUDITORIA (photo gallery bolsões)
    # ══════════════════════════════════════════════════════════════════════════

    # List of all images for selected contract
    # Keys: id, contrato, categoria, url, legenda, data_captura, autor
    audit_images: List[Dict[str, str]] = []

    # Currently open bolsão slug (e.g. "equipe", "falhas", "ferramentas", "gerais")
    audit_open_category: str = ""

    # Lightbox
    audit_lightbox_url: str = ""
    audit_lightbox_legenda: str = ""
    audit_lightbox_data: str = ""
    audit_lightbox_autor: str = ""

    # Upload dialog
    audit_show_upload: bool = False
    audit_upload_category: str = ""
    audit_upload_url: str = ""
    audit_upload_legenda: str = ""
    audit_uploading: bool = False
    audit_upload_error: str = ""

    # ══════════════════════════════════════════════════════════════════════════
    # TIMELINE
    # ══════════════════════════════════════════════════════════════════════════

    # Feed of log entries for selected contract
    # Keys: id, contrato, tipo, titulo, descricao, autor, created_at,
    #       is_document, is_cost, custo_valor, custo_categoria,
    #       anexo_url, anexo_nome
    timeline_entries: List[Dict[str, str]] = []

    # New entry form
    tl_entry_type: str = "Atualização"
    tl_titulo: str = ""
    tl_descricao: str = ""
    tl_submitting: bool = False
    tl_error: str = ""

    # mention users
    tl_mention_users: List[str] = []   # usernames disponíveis para @mention

    # custo fields
    tl_custo_valor: str = ""       # valor do custo (string para input)
    tl_custo_categoria: str = "Operacional"  # categoria do custo

    # New: file attachment
    tl_anexo_url: str = ""         # URL do arquivo no Supabase Storage após upload
    tl_anexo_nome: str = ""        # nome original do arquivo
    tl_uploading_anexo: bool = False

    # Filter + search
    tl_filter_tipo: str = ""
    tl_search: str = ""            # busca por título/descrição
    tl_search_input: str = ""  # UI-only input buffer

    # ══════════════════════════════════════════════════════════════════════════
    # MACRO/MICRO hierarchy
    # ══════════════════════════════════════════════════════════════════════════

    # Edit fields for hierarchical properties
    cron_edit_nivel: str = "macro"       # "macro" | "micro"
    cron_edit_peso: str = "100"          # peso_pct relative to parent (macro) or project (macro)
    cron_edit_parent_id: str = ""        # parent macro id (only for micros)

    # Which macro rows are expanded (showing micros) — list of macro ids
    cron_expanded_macros: List[str] = []

    # Pending activities awaiting approval (role=gestor)
    pending_activities: List[Dict[str, str]] = []
    cron_approve_loading: bool = False

    # ══════════════════════════════════════════════════════════════════════════
    # GANTT PREMIUM — weather + IA climate
    # ══════════════════════════════════════════════════════════════════════════

    # IA climate analysis result
    cron_climate_analysis: str = ""
    cron_climate_loading: bool = False

    # ══════════════════════════════════════════════════════════════════════════
    # AGENTE DE ATIVIDADES
    # ══════════════════════════════════════════════════════════════════════════

    # List of insight cards: [{type, title, body, icon, priority, atividade}]
    agente_insights: List[Dict[str, str]] = []
    agente_loading: bool = False
    agente_last_updated: str = ""  # BRT timestamp
    agente_contrato: str = ""      # contract the insights are for
    agente_error: str = ""
    agente_last_rdo_id: str = ""   # ID do último RDO que gerou os insights (cache key)

    # ──────────────────────────────────────────────────────────────────────────
    # Internal helpers
    # ──────────────────────────────────────────────────────────────────────────

    @rx.var
    def filtered_cron_rows(self) -> List[Dict[str, str]]:
        """Apply fase filter, search, and critical-only to cron_rows."""
        rows = self.cron_rows
        if self.cron_fase_filter:
            rows = [r for r in rows if r.get("fase_macro", "") == self.cron_fase_filter]
        if self.cron_show_only_critical:
            rows = [r for r in rows if r.get("critico", "") == "1"]
        if self.cron_search:
            q = self.cron_search.lower()
            rows = [
                r for r in rows
                if q in r.get("atividade", "").lower()
                or q in r.get("responsavel", "").lower()
                or q in r.get("fase", "").lower()
            ]
        return rows

    @rx.var
    def cron_unique_fases(self) -> List[str]:
        seen = []
        for r in self.cron_rows:
            f = r.get("fase_macro", "").strip()
            if f and f not in seen:
                seen.append(f)
        return seen

    @rx.var
    def cron_stats(self) -> Dict[str, str]:
        total = len(self.cron_rows)
        done = sum(1 for r in self.cron_rows if int(r.get("conclusao_pct", "0") or "0") >= 100)
        critical = sum(1 for r in self.cron_rows if r.get("critico", "") == "1")
        pct = round(done / total * 100) if total else 0
        return {
            "total": str(total),
            "done": str(done),
            "critical": str(critical),
            "pct": str(pct),
        }

    @rx.var
    def audit_category_counts(self) -> Dict[str, str]:
        counts: Dict[str, str] = {}
        for cat in AUDIT_CATEGORIES:
            slug = cat["slug"]
            counts[slug] = str(sum(1 for img in self.audit_images if img.get("categoria") == slug))
        return counts

    @rx.var
    def audit_open_images(self) -> List[Dict[str, str]]:
        if not self.audit_open_category:
            return []
        return [img for img in self.audit_images if img.get("categoria") == self.audit_open_category]

    @rx.var
    def filtered_timeline(self) -> List[Dict[str, str]]:
        rows = self.timeline_entries
        if self.tl_filter_tipo:
            rows = [e for e in rows if e.get("tipo") == self.tl_filter_tipo]
        if self.tl_search:
            q = self.tl_search.lower()
            rows = [e for e in rows if q in e.get("titulo", "").lower() or q in e.get("descricao", "").lower()]
        return rows

    @rx.var
    def audit_lightbox_open(self) -> bool:
        return self.audit_lightbox_url != ""

    @rx.var
    def cron_activity_options(self) -> List[str]:
        """List of existing activity names for dependency dropdown (excludes current edit)."""
        return [
            r.get("atividade", "")
            for r in self.cron_rows
            if r.get("atividade") and r.get("id") != self.cron_edit_id
        ]

    @rx.var
    def cron_edit_nivel_label(self) -> str:
        """Human-readable label for the current activity level."""
        labels = {"macro": "Macro (Principal)", "micro": "Micro (Sub-atividade)", "sub": "Sub (Detalhe da micro)"}
        return labels.get(self.cron_edit_nivel, self.cron_edit_nivel)

    @rx.var
    def cron_edit_parent_name(self) -> str:
        """Name of the parent activity (macro or micro) for display in the dialog."""
        if not self.cron_edit_parent_id:
            return ""
        row = next((r for r in self.cron_rows if r["id"] == self.cron_edit_parent_id), None)
        return row.get("atividade", "") if row else ""

    @rx.var
    def cron_edit_macro_name(self) -> str:
        """Name of the macro ancestor for display when creating a Sub."""
        if self.cron_edit_nivel != "sub" or not self.cron_edit_parent_id:
            return ""
        parent = next((r for r in self.cron_rows if r["id"] == self.cron_edit_parent_id), None)
        if not parent:
            return ""
        # parent is micro; find its parent (macro)
        macro_id = parent.get("parent_id", "")
        if not macro_id:
            return parent.get("fase_macro", "")
        macro = next((r for r in self.cron_rows if r["id"] == macro_id), None)
        return macro.get("atividade", "") if macro else parent.get("fase_macro", "")

    @rx.var
    def gantt_rows(self) -> List[Dict[str, str]]:
        """
        Returns filtered_cron_rows enriched with Gantt positioning data.
        Optimized: single-pass date parsing, no redundant fromisoformat calls.
        """
        from datetime import date, timedelta
        rows = self.filtered_cron_rows
        if not rows:
            return []

        # Pre-parse all dates from cron_rows in one pass to find global min/max
        today = date.today()
        global_start_d: date | None = None
        global_end_d: date | None = None
        for r in self.cron_rows:
            for key in ("inicio_iso", "termino_iso"):
                iso = r.get(key, "")
                if iso and len(iso) >= 10:
                    try:
                        d = date.fromisoformat(iso[:10])
                        if global_start_d is None or d < global_start_d:
                            global_start_d = d
                        if global_end_d is None or d > global_end_d:
                            global_end_d = d
                    except ValueError:
                        pass

        if global_start_d is None or global_end_d is None:
            return [dict(r, gantt_left_pct="0", gantt_width_pct="100", gantt_overdue="0",
                         gantt_forecast_left="", gantt_forecast_width="") for r in rows]

        global_start = global_start_d
        global_end = global_end_d
        total_days = max((global_end - global_start).days, 1)

        result = []
        for r in rows:
            s_iso = r.get("inicio_iso", "")
            e_iso = r.get("termino_iso", "")
            # Parse once per row
            s_date = global_start
            e_date = global_end
            if s_iso and len(s_iso) >= 10:
                try: s_date = date.fromisoformat(s_iso[:10])
                except ValueError: pass
            if e_iso and len(e_iso) >= 10:
                try: e_date = date.fromisoformat(e_iso[:10])
                except ValueError: pass

            left_days = (s_date - global_start).days
            dur_days = max((e_date - s_date).days, 1)
            left_pct = round(left_days / total_days * 100, 1)
            width_pct = round(dur_days / total_days * 100, 1)
            if left_pct < 0:
                left_pct = 0.0
            if left_pct + width_pct > 100:
                width_pct = 100.0 - left_pct
            overdue = "1" if e_date < today and int(r.get("conclusao_pct", "0") or "0") < 100 else "0"

            # Forecast bar — only for micros with qty data
            gantt_forecast_left = ""
            gantt_forecast_width = ""
            total_qty_f = float(r.get("total_qty", "0") or "0")
            exec_qty_f  = float(r.get("exec_qty", "0") or "0")
            dias_plan_f = int(r.get("dias_planejados", "0") or "0")
            if total_qty_f > 0 and exec_qty_f > 0 and dias_plan_f > 0:
                dias_dec = max(1, int((today - s_date).days * 5 / 7))
                prod_real_f = exec_qty_f / dias_dec
                if prod_real_f > 0:
                    saldo_f = max(0.0, total_qty_f - exec_qty_f)
                    eac_date = today + timedelta(days=int(saldo_f / prod_real_f * 1.4))
                    eac_left = round((s_date - global_start).days / total_days * 100, 1)
                    eac_w    = round(max((eac_date - s_date).days, 1) / total_days * 100, 1)
                    eac_left = max(0.0, eac_left)
                    if eac_left + eac_w > 100:
                        eac_w = 100.0 - eac_left
                    gantt_forecast_left  = str(eac_left)
                    gantt_forecast_width = str(max(eac_w, 0.5))

            # Today line position (percentage across the timeline)
            today_left = round((today - global_start).days / total_days * 100, 2)
            today_left = max(0.0, min(100.0, today_left))

            result.append(dict(
                r,
                gantt_left_pct=str(left_pct),
                gantt_width_pct=str(max(width_pct, 0.8)),
                gantt_overdue=overdue,
                gantt_forecast_left=gantt_forecast_left,
                gantt_forecast_width=gantt_forecast_width,
                gantt_today_pct=str(today_left),
            ))

        # Sort by inicio_iso chronologically (macros group with their children)
        result.sort(key=lambda x: (x.get("inicio_iso", ""), x.get("nivel", ""), x.get("atividade", "")))
        return result

    @rx.var
    def gantt_date_range(self) -> Dict[str, str]:
        """Returns {'start': 'DD/MM/YYYY', 'end': 'DD/MM/YYYY'} for display."""
        from datetime import date
        if not self.cron_rows:
            return {"start": "—", "end": "—"}
        dates_s, dates_e = [], []
        for r in self.cron_rows:
            s = r.get("inicio_iso", "")
            e = r.get("termino_iso", "")
            if s and len(s) >= 10:
                try:
                    dates_s.append(date.fromisoformat(s[:10]))
                except Exception:
                    pass
            if e and len(e) >= 10:
                try:
                    dates_e.append(date.fromisoformat(e[:10]))
                except Exception:
                    pass
        if not dates_s or not dates_e:
            return {"start": "—", "end": "—"}
        def fmt(d: date) -> str:
            return d.strftime("%d/%m/%Y")
        return {"start": fmt(min(dates_s)), "end": fmt(max(dates_e))}

    @rx.var
    def cron_macro_rows(self) -> List[Dict[str, str]]:
        """All macro-level activities (nivel=='macro' or nivel missing/empty), sorted by fase."""
        def _fase_key(r: dict) -> tuple:
            fase = str(r.get("fase", "") or "")
            parts = []
            for seg in fase.split("."):
                try:
                    parts.append(int(seg))
                except ValueError:
                    parts.append(0)
            return tuple(parts) if parts else (9999,)
        rows = [r for r in self.filtered_cron_rows if r.get("nivel", "macro") in ("macro", "")]
        return sorted(rows, key=_fase_key)

    @rx.var
    def cron_display_rows(self) -> List[Dict[str, str]]:
        """
        Flat ordered list for rx.foreach — macros interleaved with their micros.
        Each row has _display_mode: 'macro' | 'micro' | 'sub'.
        Micros only appear when their parent macro is in cron_expanded_macros.
        O(n) via pre-indexed dicts instead of nested loops.
        """
        result: List[Dict[str, str]] = []
        all_rows = self.filtered_cron_rows
        expanded = self.cron_expanded_macros

        # Fase sort key: "1" → (1,), "1.2" → (1, 2), "1.2.3" → (1, 2, 3)
        def _fase_key(r: dict) -> tuple:
            fase = str(r.get("fase", "") or "")
            parts = []
            for seg in fase.split("."):
                try:
                    parts.append(int(seg))
                except ValueError:
                    parts.append(0)
            return tuple(parts) if parts else (9999,)

        # Separate by nivel — sort each group by fase
        macros: List[Dict[str, str]] = []
        micros: List[Dict[str, str]] = []
        subs: List[Dict[str, str]] = []
        for r in all_rows:
            nivel = r.get("nivel", "macro")
            if nivel in ("macro", ""):
                macros.append(r)
            elif nivel == "micro":
                micros.append(r)
            else:
                subs.append(r)

        macros.sort(key=_fase_key)
        micros.sort(key=_fase_key)
        subs.sort(key=_fase_key)

        # Build lookup dicts — O(n) total
        micros_by_parent: Dict[str, List[Dict[str, str]]] = {}
        for m in micros:
            pid = m.get("parent_id", "")
            if pid not in micros_by_parent:
                micros_by_parent[pid] = []
            micros_by_parent[pid].append(m)

        subs_by_parent: Dict[str, List[Dict[str, str]]] = {}
        for s in subs:
            pid = s.get("parent_id", "")
            if pid not in subs_by_parent:
                subs_by_parent[pid] = []
            subs_by_parent[pid].append(s)

        macro_ids: set = {m.get("id", "") for m in macros}

        # Build id→fase lookup for dependency badge
        id_to_fase: Dict[str, str] = {r.get("id", ""): r.get("fase", "") for r in all_rows}

        def _dep_fase(row: dict) -> str:
            """Return the predecessor's fase string (e.g. '2.2'), or '' if no dependency."""
            dep_id = row.get("dependencia_id", "")
            if dep_id:
                return id_to_fase.get(dep_id, "")
            return ""

        for macro in macros:
            macro_id = macro.get("id", "")
            micro_list = micros_by_parent.get(macro_id, [])

            # For each micro child, compute its pct from subs if any — O(subs per micro)
            children: List[tuple] = []
            for micro in micro_list:
                micro_id = micro.get("id", "")
                sub_children = subs_by_parent.get(micro_id, [])
                if sub_children:
                    sub_peso = sum(int(s.get("peso_pct", "0") or "0") for s in sub_children)
                    if sub_peso > 0:
                        sub_wpct = sum(
                            int(s.get("conclusao_pct", "0") or "0") * int(s.get("peso_pct", "0") or "0")
                            for s in sub_children
                        ) / sub_peso
                    else:
                        sub_wpct = 0.0
                    micro = dict(micro, conclusao_pct=str(round(sub_wpct)))
                children.append((micro, sub_children))

            if children:
                total_peso = sum(int(m.get("peso_pct", "0") or "0") for m, _ in children)
                if total_peso > 0:
                    weighted_pct = sum(
                        int(m.get("conclusao_pct", "0") or "0") * int(m.get("peso_pct", "0") or "0")
                        for m, _ in children
                    ) / total_peso
                else:
                    weighted_pct = 0.0
                computed_pct = str(round(weighted_pct))
                has_micros = "1"
            else:
                computed_pct = macro.get("conclusao_pct", "0")
                has_micros = "0"

            is_expanded = "1" if macro_id in expanded else "0"
            result.append(dict(
                macro,
                _display_mode="macro",
                _has_micros=has_micros,
                _is_expanded=is_expanded,
                _computed_pct=computed_pct,
                _micro_count=str(len(children)),
                _dep_fase=_dep_fase(macro),
            ))

            # Append micros (and their subs) if expanded
            if macro_id in expanded:
                for micro, sub_children in children:
                    has_subs = "1" if sub_children else "0"
                    result.append(dict(
                        micro,
                        _display_mode="micro",
                        _has_micros=has_subs,
                        _is_expanded="0",
                        _computed_pct=micro.get("conclusao_pct", "0"),
                        _micro_count=str(len(sub_children)),
                        _dep_fase=_dep_fase(micro),
                    ))
                    for sub in sub_children:
                        result.append(dict(
                            sub,
                            _display_mode="sub",
                            _has_micros="0",
                            _is_expanded="0",
                            _computed_pct=sub.get("conclusao_pct", "0"),
                            _micro_count="0",
                            _dep_fase=_dep_fase(sub),
                        ))

        # Orphan micros (no macro parent in current filter)
        for micro in micros:
            if micro.get("parent_id", "") not in macro_ids:
                result.append(dict(
                    micro,
                    _display_mode="micro",
                    _has_micros="0",
                    _is_expanded="0",
                    _computed_pct=micro.get("conclusao_pct", "0"),
                    _micro_count="0",
                    _dep_fase=_dep_fase(micro),
                ))

        return result

    @rx.var
    def cron_pending_rows(self) -> List[Dict[str, str]]:
        """Activities flagged as pendente_aprovacao=True."""
        return [r for r in self.cron_rows if r.get("pendente_aprovacao", "0") == "1"]

    @rx.var
    def cron_parent_options(self) -> List[Dict[str, str]]:
        """List of macros for parent dropdown when creating/editing a micro."""
        return [
            {"id": r.get("id", ""), "label": r.get("atividade", "")}
            for r in self.cron_rows
            if r.get("nivel", "macro") in ("macro", "") and r.get("id") != self.cron_edit_id
        ]

    @rx.var
    def cron_dep_options(self) -> List[Dict[str, str]]:
        """All activities as dependency targets, ordered hierarchically (macro → micro → sub).
        Label uses level prefix to show hierarchy: macro=bold, micro=↳, sub=↳↳"""
        def _fase_key(r: dict) -> tuple:
            fase = str(r.get("fase", "") or "")
            parts = []
            for seg in fase.split("."):
                try:
                    parts.append(int(seg))
                except ValueError:
                    parts.append(0)
            return tuple(parts) if parts else (9999,)

        edit_id = self.cron_edit_id
        rows = [r for r in self.cron_rows if r.get("id") != edit_id]

        # Separate and sort each level
        macros = sorted([r for r in rows if r.get("nivel", "macro") in ("macro", "")], key=_fase_key)
        micros_all = sorted([r for r in rows if r.get("nivel") == "micro"], key=_fase_key)
        subs_all = sorted([r for r in rows if r.get("nivel") == "sub"], key=_fase_key)

        # Index children by parent
        micros_by_parent: dict = {}
        for m in micros_all:
            pid = m.get("parent_id", "")
            micros_by_parent.setdefault(pid, []).append(m)

        subs_by_parent: dict = {}
        for s in subs_all:
            pid = s.get("parent_id", "")
            subs_by_parent.setdefault(pid, []).append(s)

        result = []
        for macro in macros:
            macro_id = macro.get("id", "")
            fase = macro.get("fase", "")
            nome = macro.get("atividade", "")
            result.append({"id": macro_id, "label": f"{fase} {nome}".strip()})
            for micro in micros_by_parent.get(macro_id, []):
                micro_id = micro.get("id", "")
                m_fase = micro.get("fase", "")
                m_nome = micro.get("atividade", "")
                result.append({"id": micro_id, "label": f"  ↳ {m_fase} {m_nome}".strip()})
                for sub in subs_by_parent.get(micro_id, []):
                    s_fase = sub.get("fase", "")
                    s_nome = sub.get("atividade", "")
                    result.append({"id": sub.get("id", ""), "label": f"    ↳↳ {s_fase} {s_nome}".strip()})
        return result

    @rx.var
    def cron_micro_options(self) -> List[Dict[str, str]]:
        """List of micro activities available as parent for sub-activities.
        Excludes the activity being edited."""
        return [
            {"id": r.get("id", ""), "label": r.get("fase", "") + " " + r.get("atividade", "")}
            for r in self.cron_rows
            if r.get("nivel", "macro") == "micro" and r.get("id") != self.cron_edit_id
        ]

    # ── Forecast / Produtividade computed vars ────────────────────────────────

    @rx.var
    def cron_forecast_rows(self) -> List[Dict[str, str]]:
        """
        Pass-through para o cache pré-computado em background.
        O cálculo pesado (O(n) com parsing de datas) roda em _compute_forecast_rows()
        via run_in_executor dentro de load_cronograma — nunca no event loop.
        """
        return self._cron_forecast_cache

    @rx.var
    def cron_forecast_filtered(self) -> List[Dict[str, str]]:
        """
        Cache filtrado + ordenado. Subs sempre ficam coladas à sua micro (não filtradas
        individualmente). Filtra apenas micros; subs seguem junto.
        """
        rows = self._cron_forecast_cache
        f = self.cron_forecast_filter
        _order = {"abaixo": 0, "dentro": 1, "acima": 2, "sem_dados": 3, "concluida": 4}

        # Separar micros de subs
        micros = [r for r in rows if r.get("_is_sub", "0") != "1"]
        subs_by_micro: dict = {}
        for r in rows:
            if r.get("_is_sub", "0") == "1":
                pid = r.get("parent_id", "")
                subs_by_micro.setdefault(pid, []).append(r)

        # Filtrar micros conforme filtro ativo
        if f == "execucao":
            filtered_micros = [r for r in micros
                                if r.get("_tendencia") != "concluida"
                                and r.get("_dia_atual", "0") != "0"]
        elif f == "concluida":
            filtered_micros = [r for r in micros if r.get("_tendencia") == "concluida"]
        elif f == "prevista":
            filtered_micros = [r for r in micros if r.get("_dia_atual", "0") == "0"]
        else:
            filtered_micros = list(micros)

        # Ordenar micros por desvio
        sorted_micros = sorted(filtered_micros,
                               key=lambda r: _order.get(r.get("_tendencia", "sem_dados"), 99))

        # Reinserir subs logo abaixo de cada micro
        result = []
        for micro in sorted_micros:
            result.append(micro)
            for sub in subs_by_micro.get(micro.get("id", ""), []):
                result.append(sub)
        return result

    @rx.var
    def cron_forecast_kpis(self) -> Dict[str, str]:
        """
        KPIs do summary bar — single-pass, ignora subs nas contagens.
        """
        n_exec = n_conc = n_prev = n_risco = 0
        desvios: list = []
        for r in self._cron_forecast_cache:
            if r.get("_is_sub", "0") == "1":
                continue  # subs não contam nos KPIs
            t   = r.get("_tendencia", "sem_dados")
            dia = r.get("_dia_atual", "0")
            if t == "concluida":
                n_conc += 1
            elif dia == "0":
                n_prev += 1
            else:
                n_exec += 1
                if t == "abaixo":
                    n_risco += 1
                if t not in ("sem_dados",):
                    try:
                        desvios.append(float(r.get("_desvio_pct", "0")))
                    except ValueError:
                        pass
        desvio_medio = round(sum(desvios) / len(desvios), 1) if desvios else 0.0
        return {
            "em_exec":         str(n_exec),
            "concluidas":      str(n_conc),
            "previstas":       str(n_prev),
            "em_risco":        str(n_risco),
            "desvio_medio":    f"{desvio_medio:+.1f}%",
            "desvio_positivo": "1" if desvio_medio >= 0 else "0",
        }

    @rx.var
    def cron_kpi_dashboard(self) -> Dict[str, str]:
        """
        KPIs de alto nível do cronograma — single-pass otimizado.
        Evita reprocessar cron_forecast_rows (que é uma computed var pesada separada).
        """
        from datetime import date
        today = date.today()
        tol = 10.0

        micros = [r for r in self.cron_rows if r.get("nivel", "") == "micro"]
        n = len(micros)
        if not n:
            return {
                "pct_fisico_programado_hoje": "0",
                "pct_fisico_realizado": "0",
                "desvio_pp": "0",
                "atividades_em_risco": "0",
                "atividades_atrasadas": "0",
                "atividades_adiantadas": "0",
                "producao_total_prevista": "0",
                "producao_total_realizada": "0",
                "total_micros": "0",
            }

        # Single pass — compute all KPIs together
        total_peso = 0
        peso_pct_sum = 0.0
        pct_sum = 0
        vencidas = 0
        atrasadas = 0
        prod_prev = 0.0
        prod_real_qty = 0.0
        em_risco = 0
        adiantadas = 0

        for r in micros:
            peso  = int(r.get("peso_pct", "0") or "0")
            pct   = int(r.get("conclusao_pct", "0") or "0")
            t_iso = r.get("termino_iso", "")
            tqty  = float(r.get("total_qty", "0") or "0")
            eqty  = float(r.get("exec_qty", "0") or "0")
            dias  = int(r.get("dias_planejados", "0") or "0")

            total_peso    += peso
            peso_pct_sum  += peso * pct
            pct_sum       += pct
            prod_prev     += tqty
            prod_real_qty += eqty

            # Date checks — parse once per row
            if t_iso and len(t_iso) >= 10:
                try:
                    t_date = date.fromisoformat(t_iso[:10])
                    if t_date <= today:
                        vencidas += 1
                    if t_date < today and pct < 100:
                        atrasadas += 1
                except ValueError:
                    pass

            # Inline forecast tendency (same logic as cron_forecast_rows)
            s_iso = r.get("inicio_iso", "")
            if eqty > 0 and tqty > 0 and dias > 0 and s_iso and len(s_iso) >= 10:
                try:
                    d_inicio = date.fromisoformat(s_iso[:10])
                    dias_dec = max(0, (today - d_inicio).days) if d_inicio <= today else 0
                    if dias_dec >= 3:
                        prod_plan = tqty / dias
                        prod_r    = eqty / max(1, int(dias_dec * 5 / 7))
                        desvio    = (prod_r - prod_plan) / prod_plan * 100 if prod_plan > 0 else 0.0
                        if desvio <= -tol and pct < 100:
                            em_risco += 1
                        elif desvio >= tol:
                            adiantadas += 1
                except ValueError:
                    pass

        pct_realizado  = peso_pct_sum / total_peso if total_peso > 0 else (pct_sum / n)
        pct_programado = vencidas / n * 100
        desvio_pp      = round(pct_realizado - pct_programado, 1)

        return {
            "pct_fisico_programado_hoje": str(round(pct_programado, 1)),
            "pct_fisico_realizado": str(round(pct_realizado, 1)),
            "desvio_pp": f"{desvio_pp:+.1f}",
            "atividades_em_risco": str(em_risco),
            "atividades_atrasadas": str(atrasadas),
            "atividades_adiantadas": str(adiantadas),
            "producao_total_prevista": str(round(prod_prev, 1)),
            "producao_total_realizada": str(round(prod_real_qty, 1)),
            "total_micros": str(n),
        }

    @rx.var
    def cron_kpi_popup_rows(self) -> List[Dict[str, str]]:
        """
        Linhas detalhadas para o popup de KPI clicável.
        Filtra micro-atividades de acordo com cron_kpi_popup:
          'programado' → término <= hoje (deveriam estar concluídas ou em execução hoje)
          'atrasadas'  → término < hoje e pct < 100
          'em_risco'   → tendência de produção ≤ -10%
          'adiantadas' → tendência de produção ≥ +10%
          'realizado'  → todas com pct > 0 (concluídas ou em andamento)
        """
        from datetime import date
        mode = self.cron_kpi_popup
        if not mode:
            return []

        today = date.today()
        tol = 10.0
        micros = [r for r in self.cron_rows if r.get("nivel", "") == "micro"]
        result = []

        for r in micros:
            pct   = int(r.get("conclusao_pct", "0") or "0")
            t_iso = r.get("termino_iso", "")
            s_iso = r.get("inicio_iso", "")
            tqty  = float(r.get("total_qty", "0") or "0")
            eqty  = float(r.get("exec_qty", "0") or "0")
            dias  = int(r.get("dias_planejados", "0") or "0")
            ativ  = r.get("atividade", "")
            fase  = r.get("fase_macro", "")
            resp  = r.get("responsavel", "")
            unid  = r.get("unidade", "")

            include = False
            desvio_str = ""
            saldo_str = ""

            if mode == "programado":
                # Programadas para hoje: início <= hoje <= término
                if t_iso and s_iso and len(t_iso) >= 10 and len(s_iso) >= 10:
                    try:
                        s_d = date.fromisoformat(s_iso[:10])
                        t_d = date.fromisoformat(t_iso[:10])
                        if s_d <= today <= t_d and pct < 100:
                            include = True
                            saldo_str = f"{max(0.0, tqty - eqty):.1f} {unid}".strip() if tqty > 0 else ""
                    except ValueError:
                        pass
                # Também inclui atrasadas (termino < hoje, pct < 100) — atraso acumulado
                if not include and t_iso and len(t_iso) >= 10:
                    try:
                        t_d = date.fromisoformat(t_iso[:10])
                        if t_d < today and pct < 100:
                            include = True
                            dias_atraso = (today - t_d).days
                            saldo_str = f"{max(0.0, tqty - eqty):.1f} {unid} · {dias_atraso}d atraso".strip() if tqty > 0 else f"{dias_atraso}d atraso"
                    except ValueError:
                        pass

            elif mode == "atrasadas":
                if t_iso and len(t_iso) >= 10:
                    try:
                        t_d = date.fromisoformat(t_iso[:10])
                        if t_d < today and pct < 100:
                            include = True
                            dias_atraso = (today - t_d).days
                            saldo_str = f"{dias_atraso}d vencida"
                    except ValueError:
                        pass

            elif mode == "em_risco":
                if eqty > 0 and tqty > 0 and dias > 0 and s_iso and len(s_iso) >= 10:
                    try:
                        d_inicio = date.fromisoformat(s_iso[:10])
                        dias_dec = max(0, (today - d_inicio).days) if d_inicio <= today else 0
                        if dias_dec >= 3:
                            prod_plan = tqty / dias
                            prod_r    = eqty / max(1, int(dias_dec * 5 / 7))
                            desvio    = (prod_r - prod_plan) / prod_plan * 100 if prod_plan > 0 else 0.0
                            if desvio <= -tol and pct < 100:
                                include = True
                                desvio_str = f"{desvio:+.1f}%"
                                saldo_str = f"{max(0.0, tqty - eqty):.1f} {unid} restantes".strip()
                    except ValueError:
                        pass

            elif mode == "adiantadas":
                if eqty > 0 and tqty > 0 and dias > 0 and s_iso and len(s_iso) >= 10:
                    try:
                        d_inicio = date.fromisoformat(s_iso[:10])
                        dias_dec = max(0, (today - d_inicio).days) if d_inicio <= today else 0
                        if dias_dec >= 3:
                            prod_plan = tqty / dias
                            prod_r    = eqty / max(1, int(dias_dec * 5 / 7))
                            desvio    = (prod_r - prod_plan) / prod_plan * 100 if prod_plan > 0 else 0.0
                            if desvio >= tol:
                                include = True
                                desvio_str = f"+{desvio:.1f}%"
                    except ValueError:
                        pass

            elif mode == "realizado":
                if pct > 0:
                    include = True
                    saldo_str = f"{eqty:.1f}/{tqty:.1f} {unid}".strip() if tqty > 0 else ""

            if include:
                dep = r.get("dependencia", "") or ""
                result.append({
                    "atividade":  ativ,
                    "fase_macro": fase,
                    "responsavel": resp,
                    "conclusao_pct": str(pct),
                    "desvio": desvio_str,
                    "saldo": saldo_str,
                    "unidade": unid,
                    "termino_iso": t_iso,
                    "dependencia": dep,
                })

        # Sort: atrasadas primeiro, depois por término
        result.sort(key=lambda x: x.get("termino_iso", ""))
        return result

    # ══════════════════════════════════════════════════════════════════════════
    # CRONOGRAMA — Load & CRUD
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def load_cronograma(self, contrato: str):
        """Load activities for a given contract from hub_atividades table."""
        async with self:
            self.cron_loading = True
            self.cron_rows = []
            self.cron_fase_filter = ""
            self.cron_search = ""

        normalized: list = []
        dias_uteis_str = "seg,ter,qua,qui,sex"
        _prev_contrato = contrato

        try:
            # ── 1. Load rows from DB ───────────────────────────────────────────
            try:
                contrato_rows = sb_select("contratos", filters={"contrato": contrato}, limit=1)
                if contrato_rows:
                    dias_uteis_str = str(contrato_rows[0].get("dias_uteis_semana", "") or "seg,ter,qua,qui,sex")

                rows = sb_select(
                    "hub_atividades",
                    filters={"contrato": contrato},
                    order="fase_macro.asc,inicio_previsto.asc",
                    limit=500,
                )
                for r in rows:
                    fase = _norm_str(r.get("fase", ""))
                    pendente_raw = r.get("pendente_aprovacao", False)
                    pendente = "1" if str(pendente_raw or "").upper() in ("TRUE", "1", "SIM", "YES") else "0"
                    normalized.append({
                        "id":              _norm_str(r.get("id")),
                        "contrato":        _norm_str(r.get("contrato")),
                        "fase_macro":      _norm_str(r.get("fase_macro")),
                        "fase":            fase,
                        "atividade":       _norm_str(r.get("atividade")),
                        "responsavel":     _norm_str(r.get("responsavel"), "—"),
                        "inicio_previsto": _utc_date_to_br(_norm_str(r.get("inicio_previsto"))),
                        "termino_previsto": _utc_date_to_br(_norm_str(r.get("termino_previsto"))),
                        # raw ISO dates kept for Gantt/weather lookups
                        "inicio_iso": _norm_str(r.get("inicio_previsto"))[:10],
                        "termino_iso": _norm_str(r.get("termino_previsto"))[:10],
                        "conclusao_pct":   _norm_pct(r.get("conclusao_pct")),
                        "critico":         "1" if str(r.get("critico", "") or "").upper() in ("TRUE", "1", "SIM") else "0",
                        "dependencia":     _norm_str(r.get("dependencia")),
                        "observacoes":     _norm_str(r.get("observacoes")),
                        "color":           _fase_color(fase),
                        # Hierarchical fields
                        "nivel":           _norm_str(r.get("nivel"), "macro"),
                        "parent_id":       _norm_str(r.get("parent_id")),
                        "peso_pct":        _norm_pct(r.get("peso_pct") if r.get("peso_pct") is not None else 100),
                        "pendente_aprovacao": pendente,
                        # Quantity tracking
                        "total_qty":       _norm_str(r.get("total_qty", "0") or "0"),
                        "exec_qty":        _norm_str(r.get("exec_qty", "0") or "0"),
                        "unidade":         _norm_str(r.get("unidade", "")),
                        "dias_planejados": _norm_str(r.get("dias_planejados", "0") or "0"),
                        "dependencia_id":  _norm_str(r.get("dependencia_id", "")),
                        # dep_tipo: fallback for legacy rows — if dependencia_id exists → 'tradicional'
                        "dep_tipo": _norm_str(
                            r.get("dep_tipo")
                            or ("tradicional" if r.get("dependencia_id") else "sem_dep")
                        ),
                        # Forecast fields (new)
                        "status_atividade": _norm_str(r.get("status_atividade", "nao_iniciada") or "nao_iniciada"),
                        "tipo_medicao":     _norm_str(r.get("tipo_medicao", "quantidade") or "quantidade"),
                        "frente_servico":   _norm_str(r.get("frente_servico", "")),
                        "data_inicio_real": _utc_date_to_br(_norm_str(r.get("data_inicio_real") or "")),
                        "data_fim_real":    _utc_date_to_br(_norm_str(r.get("data_fim_real") or "")),
                        "data_fim_prevista": _utc_date_to_br(_norm_str(r.get("data_fim_prevista") or "")),
                        "data_inicio_real_iso": _norm_str(r.get("data_inicio_real") or "")[:10],
                        "data_fim_real_iso":    _norm_str(r.get("data_fim_real") or "")[:10],
                        "efetivo_alocado":      _norm_str(r.get("efetivo_alocado", "0") or "0"),
                        # Âncora temporal: data do último RDO que atualizou esta atividade
                        "last_rdo_date":        _norm_str(r.get("last_rdo_date") or "")[:10],
                    })
            except Exception as e:
                logger.error(f"load_cronograma (db fetch): {e}", exc_info=True)

            # ── 2. Pré-computa forecast fora do lock (CPU-bound) ───────────────
            import asyncio as _aio_fc
            _fc_loop = _aio_fc.get_running_loop()
            forecast_cache = await _fc_loop.run_in_executor(
                None, lambda: _compute_forecast_rows(normalized)
            )

            async with self:
                self.cron_rows = normalized
                self._cron_forecast_cache = forecast_cache
                self.cron_working_days_str = dias_uteis_str
                _prev_contrato = self.cron_contrato
                self.cron_contrato = contrato

            # ── Update projetos_list (reactive var) with fresh progress from Supabase ──
            if normalized and contrato:
                progress_map = {
                    r["id"]: {"conclusao_pct": r.get("conclusao_pct", "0"), "peso_pct": r.get("peso_pct", "100")}
                    for r in normalized if r.get("id")
                }
                from bomtempo.state.global_state import GlobalState as _GS
                yield _GS.update_projetos_list_progress(contrato, progress_map)

            # If switching contracts, load persisted insights from Supabase for the new one
            if contrato and contrato != _prev_contrato:
                yield HubState.load_persisted_insights(contrato)

        except Exception as e:
            logger.error(f"load_cronograma (forecast/state): {e}", exc_info=True)
            async with self:
                self.cron_rows = normalized

        finally:
            async with self:
                self.cron_loading = False

    def set_cron_fase_filter(self, value: str):
        self.cron_fase_filter = "" if self.cron_fase_filter == value else value

    def set_cron_search(self, value: str):
        self.cron_search = value

    def commit_cron_search(self, _value: str = ""):
        """Commit search from blur or Enter key — only then triggers filtered_cron_rows recalc."""
        self.cron_search = self.cron_search_input

    def set_cron_search_input(self, value: str):
        """Update local input var without triggering filtered_cron_rows recalc."""
        self.cron_search_input = value

    def handle_cron_search_key(self, key: str):
        if key == "Enter":
            self.cron_search = self.cron_search_input

    def toggle_cron_critical(self):
        self.cron_show_only_critical = not self.cron_show_only_critical

    @rx.event(background=True)
    async def recalculate_cron_dates(self):
        """Recalculate termino_previsto for all activities that have inicio_previsto + dias_planejados,
        using the current project's working days config. Useful after changing dias_uteis_semana."""
        from bomtempo.core.supabase_client import sb_update

        contrato = ""
        working_days_str = "seg,ter,qua,qui,sex"
        rows_snapshot = []
        async with self:
            contrato = self.cron_contrato
            working_days_str = self.cron_working_days_str
            rows_snapshot = list(self.cron_rows)

        if not contrato or not rows_snapshot:
            yield rx.toast.warning("Nenhuma atividade carregada.", duration=3000)
            return

        wd = _parse_dias_uteis(working_days_str)
        updated = 0
        for r in rows_snapshot:
            inicio_iso = r.get("inicio_iso", "")
            dias_raw = r.get("dias_planejados", "0")
            row_id = r.get("id", "")
            if not inicio_iso or not row_id:
                continue
            try:
                dias = int(dias_raw or 0)
            except (ValueError, TypeError):
                dias = 0
            if dias <= 0:
                continue
            new_termino = _add_working_days(inicio_iso, dias, wd)
            try:
                sb_update("hub_atividades", {"id": row_id}, {"termino_previsto": new_termino})
                updated += 1
            except Exception as ex:
                logger.warning(f"recalculate_cron_dates: erro ao atualizar {row_id}: {ex}")

        if updated:
            yield rx.toast.success(f"{updated} datas de término recalculadas.", duration=4000)
            yield HubState.load_cronograma(contrato)
        else:
            yield rx.toast.info("Nenhuma atividade com início + dias para recalcular.", duration=3000)

    # ── Dialog open/close ─────────────────────────────────────────────────────

    def open_cron_new_root(self):
        self.open_cron_new("")

    def open_cron_new(self, parent_id: str = ""):
        self.cron_edit_id = ""
        self.cron_edit_atividade = ""
        self.cron_edit_fase_macro = ""
        self.cron_edit_fase = ""
        self.cron_edit_responsavel = ""
        self.cron_edit_inicio = ""
        self.cron_edit_termino = ""
        self.cron_edit_pct = "0"
        self.cron_edit_critico = False
        self.cron_edit_dependencia = ""
        self.cron_edit_dependencia_id = ""
        self.cron_edit_dep_tipo = "sem_dep"
        self.cron_edit_observacoes = ""
        self.cron_edit_total_qty = ""
        self.cron_edit_unidade = ""
        self.cron_edit_dias_planejados = ""
        self.cron_edit_status_atividade = "nao_iniciada"
        self.cron_edit_tipo_medicao = "quantidade"
        self.cron_edit_efetivo_alocado = ""
        self.cron_error = ""
        # Hierarchy defaults — auto-detect nivel based on parent's nivel
        if parent_id:
            self.cron_edit_parent_id = parent_id
            parent = next((r for r in self.cron_rows if r["id"] == parent_id), None)
            if parent:
                parent_nivel = parent.get("nivel", "macro")
                # micro under macro → sub under micro
                self.cron_edit_nivel = "sub" if parent_nivel == "micro" else "micro"
                self.cron_edit_fase_macro = parent.get("fase_macro", "")
                parent_fase = parent.get("fase", "")
                if parent_fase:
                    child_nivel = self.cron_edit_nivel  # "micro" or "sub"
                    siblings = [
                        r for r in self.cron_rows
                        if r.get("parent_id") == parent_id and r.get("nivel") == child_nivel
                    ]
                    self.cron_edit_fase = f"{parent_fase}.{len(siblings) + 1}"
                else:
                    self.cron_edit_fase = ""
            else:
                self.cron_edit_nivel = "micro"
        else:
            self.cron_edit_nivel = "macro"
            self.cron_edit_parent_id = ""
            # Auto-number macro: next integer after the highest macro fase
            # (use any existing row's contrato — all cron_rows share the same selected contract)
            macro_siblings = [
                r for r in self.cron_rows
                if r.get("nivel", "macro") in ("macro", "")
            ]
            if macro_siblings:
                max_n = 0
                for r in macro_siblings:
                    try:
                        n = int(str(r.get("fase", "0") or "0").split(".")[0])
                        if n > max_n:
                            max_n = n
                    except ValueError:
                        pass
                self.cron_edit_fase = str(max_n + 1)
            else:
                self.cron_edit_fase = "1"
        self.cron_edit_peso = "100"
        self.cron_show_dialog = True

    def open_cron_edit(self, row_id: str):
        row = next((r for r in self.cron_rows if r["id"] == row_id), None)
        if not row:
            return
        self.cron_edit_id = row_id
        self.cron_edit_atividade = row.get("atividade", "")
        self.cron_edit_fase_macro = row.get("fase_macro", "")
        self.cron_edit_fase = row.get("fase", "")
        self.cron_edit_responsavel = row.get("responsavel", "")
        self.cron_edit_inicio = row.get("inicio_iso", row.get("inicio_previsto", ""))
        self.cron_edit_termino = row.get("termino_iso", row.get("termino_previsto", ""))
        self.cron_edit_pct = row.get("conclusao_pct", "0")
        self.cron_edit_critico = row.get("critico", "0") == "1"
        self.cron_edit_dependencia = row.get("dependencia", "")
        self.cron_edit_dependencia_id = row.get("dependencia_id", "")
        self.cron_edit_dep_tipo = row.get("dep_tipo", "sem_dep") or "sem_dep"
        self.cron_edit_observacoes = row.get("observacoes", "")
        self.cron_edit_total_qty = row.get("total_qty", "")
        self.cron_edit_unidade = row.get("unidade", "")
        self.cron_edit_dias_planejados = row.get("dias_planejados", "")
        self.cron_edit_status_atividade = row.get("status_atividade", "nao_iniciada") or "nao_iniciada"
        self.cron_edit_tipo_medicao = row.get("tipo_medicao", "quantidade") or "quantidade"
        self.cron_edit_efetivo_alocado = str(row.get("efetivo_alocado", "") or "")
        # Hierarchy
        self.cron_edit_nivel = row.get("nivel", "macro")
        self.cron_edit_parent_id = row.get("parent_id", "")
        self.cron_edit_peso = str(row.get("peso_pct", "100") or "100")
        self.cron_error = ""
        self.cron_show_dialog = True

    def open_pending_review(self, row_id: str):
        """Open the edit dialog pre-filled for a pending-approval activity."""
        self.open_cron_edit(row_id)
        self.cron_pending_review_id = row_id

    def close_cron_dialog(self):
        self.cron_show_dialog = False
        self.cron_pending_review_id = ""

    def set_cron_show_dialog(self, v: bool):
        self.cron_show_dialog = v
        if not v:
            self.cron_pending_review_id = ""

    def set_cron_edit_atividade(self, v: str): self.cron_edit_atividade = v
    def set_cron_edit_fase_macro(self, v: str):
        # Para micro/sub a fase_macro é herdada do pai — bloqueado para edição.
        # A UI já deve renderizar o campo somente leitura, mas este guard evita
        # alterações acidentais via chamadas diretas ou bugs no frontend.
        if self.cron_edit_nivel in ("micro", "sub"):
            return
        self.cron_edit_fase_macro = v
        # Para atividades macro, o nome é a própria fase macro
        if self.cron_edit_nivel == "macro":
            self.cron_edit_atividade = v
    def set_cron_edit_nivel(self, v: str):
        self.cron_edit_nivel = v
        # Ao mudar para macro, sincroniza o nome com fase_macro
        if v == "macro" and self.cron_edit_fase_macro.strip():
            self.cron_edit_atividade = self.cron_edit_fase_macro
        # Ao mudar para sub ou micro, limpa o parent_id para forçar seleção
        if v in ("micro", "sub"):
            self.cron_edit_parent_id = ""
    def set_cron_edit_fase(self, v: str): self.cron_edit_fase = v
    def set_cron_edit_responsavel(self, v: str): self.cron_edit_responsavel = v
    def set_cron_edit_inicio(self, v: str):
        self.cron_edit_inicio = v
        # Auto-recalculate termino if dias_planejados is set
        if self.cron_edit_dias_planejados.strip():
            try:
                dias = int(self.cron_edit_dias_planejados.strip())
                wd = _parse_dias_uteis(self.cron_working_days_str)
                self.cron_edit_termino = _add_working_days(v, dias, wd)
            except Exception:
                pass
    def set_cron_edit_termino(self, v: str):
        self.cron_edit_termino = v
        # Ao preencher termino manualmente → recalcular dias_planejados (dias úteis)
        if v and self.cron_edit_inicio:
            try:
                wd = _parse_dias_uteis(self.cron_working_days_str)
                dias = _count_working_days(self.cron_edit_inicio, v, wd)
                if dias > 0:
                    self.cron_edit_dias_planejados = str(dias)
            except Exception:
                pass
    def set_cron_edit_pct(self, v):
        """Set conclusao_pct. For 'progresso' dependency type, cap at predecessor's current pct."""
        try:
            pct = max(0, min(100, int(float(str(v) or "0"))))
        except (ValueError, TypeError):
            self.cron_edit_pct = str(v)  # keep raw while user is typing
            return
        if self.cron_edit_dep_tipo == "progresso" and self.cron_edit_dependencia_id:
            pred = next(
                (r for r in self.cron_rows if r["id"] == self.cron_edit_dependencia_id),
                None,
            )
            if pred:
                pred_pct = int(pred.get("conclusao_pct", "0") or "0")
                if pct > pred_pct:
                    self.cron_error = (
                        f"⚠ Avanço limitado pelo progresso da predecessora: "
                        f"máximo {pred_pct}% (predecessora em {pred_pct}%)."
                    )
                    pct = pred_pct
                else:
                    self.cron_error = ""
        self.cron_edit_pct = str(pct)
    def toggle_cron_edit_critico(self): self.cron_edit_critico = not self.cron_edit_critico
    def set_cron_edit_dep_tipo(self, v: str):
        """Switch dependency mode. Clears predecessor selection when moving to 'sem_dep'."""
        self.cron_edit_dep_tipo = v
        if v == "sem_dep":
            self.cron_edit_dependencia_id = ""
            self.cron_edit_dependencia = ""

    def set_cron_edit_dependencia(self, v: str): self.cron_edit_dependencia = "" if v == "__none__" else v
    def set_cron_edit_dependencia_id(self, dep_id: str):
        """Select a dependency by activity id — auto-fill inicio from its termino."""
        if dep_id in ("", "__none__"):
            self.cron_edit_dependencia_id = ""
            self.cron_edit_dependencia = ""
            return
        self.cron_edit_dependencia_id = dep_id
        dep_row = next((r for r in self.cron_rows if r["id"] == dep_id), None)
        if dep_row:
            self.cron_edit_dependencia = dep_row.get("atividade", "")
            dep_termino = dep_row.get("termino_iso", "")
            if dep_termino:
                self.cron_edit_inicio = dep_termino
                # Recalculate termino if dias_planejados is set
                if self.cron_edit_dias_planejados.strip():
                    try:
                        dias = int(self.cron_edit_dias_planejados.strip())
                        wd = _parse_dias_uteis(self.cron_working_days_str)
                        self.cron_edit_termino = _add_working_days(dep_termino, dias, wd)
                    except Exception:
                        pass
    def set_cron_edit_dias_planejados(self, v):
        v = str(v) if v is not None else ""
        self.cron_edit_dias_planejados = v
        # Auto-calculate termino when dias changes and inicio is set
        if v.strip() and self.cron_edit_inicio:
            try:
                dias = int(v.strip())
                wd = _parse_dias_uteis(self.cron_working_days_str)
                self.cron_edit_termino = _add_working_days(self.cron_edit_inicio, dias, wd)
            except Exception:
                pass
    def set_cron_edit_total_qty(self, v): self.cron_edit_total_qty = str(v) if v is not None else ""
    def set_cron_edit_unidade(self, v: str): self.cron_edit_unidade = v
    def set_cron_edit_observacoes(self, v: str): self.cron_edit_observacoes = v
    def set_cron_edit_status_atividade(self, v: str): self.cron_edit_status_atividade = v
    def set_cron_edit_tipo_medicao(self, v: str):
        self.cron_edit_tipo_medicao = v
        # Reset quantity/unit fields when switching to non-quantity modes
        if v == "marco":
            self.cron_edit_total_qty = "1"
            self.cron_edit_unidade = "marco"
        elif v == "percentual":
            self.cron_edit_total_qty = "0"
            self.cron_edit_unidade = ""
    def set_cron_edit_peso(self, v): self.cron_edit_peso = str(v) if v is not None else "100"
    def set_cron_edit_efetivo_alocado(self, v): self.cron_edit_efetivo_alocado = str(v) if v is not None else ""
    def set_cron_edit_parent_id(self, v: str):
        self.cron_edit_parent_id = v
        if not v:
            return
        parent = next((r for r in self.cron_rows if r["id"] == v), None)
        if not parent:
            return
        # Inherit fase_macro from parent (works for both micro→macro and sub→micro)
        self.cron_edit_fase_macro = parent.get("fase_macro", "")
        parent_fase = parent.get("fase", "")
        # Auto-number: count existing siblings of same nivel → next index
        if parent_fase:
            child_nivel = self.cron_edit_nivel  # "micro" or "sub"
            siblings = [
                r for r in self.cron_rows
                if r.get("parent_id") == v and r.get("nivel") == child_nivel
            ]
            self.cron_edit_fase = f"{parent_fase}.{len(siblings) + 1}"
        else:
            self.cron_edit_fase = ""

    def toggle_macro_expanded(self, macro_id: str):
        if macro_id in self.cron_expanded_macros:
            self.cron_expanded_macros = [x for x in self.cron_expanded_macros if x != macro_id]
        else:
            new_list = list(self.cron_expanded_macros)
            new_list.append(macro_id)
            self.cron_expanded_macros = new_list

    # ── Save ─────────────────────────────────────────────────────────────────

    @rx.event(background=True)
    async def save_cron_activity(self):
        """INSERT or UPDATE a hub_atividade row."""
        from bomtempo.state.global_state import GlobalState

        contrato = ""
        atividade_nome = ""
        edit_id = ""
        old_inicio = ""
        old_termino = ""
        edit_inicio = ""
        edit_termino = ""
        edit_pct = 0
        edit_critico = False
        edit_dependencia = ""
        edit_dependencia_id = ""
        edit_observacoes = ""
        edit_fase_macro = ""
        edit_fase = ""
        edit_responsavel = ""
        edit_nivel = "macro"
        edit_parent_id = ""
        edit_peso = 100
        edit_total_qty = 0.0
        edit_unidade = ""
        edit_dias_planejados = 0
        edit_status_atividade = "nao_iniciada"
        edit_tipo_medicao = "quantidade"
        pending_review_id = ""
        username = ""

        async with self:
            if not self.cron_edit_atividade.strip():
                self.cron_error = "Nome da atividade é obrigatório."
                return
            self.cron_saving = True
            self.cron_error = ""
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
            atividade_nome = self.cron_edit_atividade.strip()
            edit_id = self.cron_edit_id
            pending_review_id = self.cron_pending_review_id
            edit_fase_macro = self.cron_edit_fase_macro.strip()
            edit_fase = self.cron_edit_fase.strip()
            edit_responsavel = self.cron_edit_responsavel.strip()
            edit_inicio = self.cron_edit_inicio or None
            edit_termino = self.cron_edit_termino or None
            edit_pct = int(self.cron_edit_pct or 0)
            edit_critico = self.cron_edit_critico
            edit_dep_tipo = self.cron_edit_dep_tipo or "sem_dep"
            # Re-enforce progresso cap on save (not just on UI input)
            if edit_dep_tipo == "progresso" and self.cron_edit_dependencia_id:
                pred = next(
                    (r for r in self.cron_rows if r["id"] == self.cron_edit_dependencia_id),
                    None,
                )
                if pred:
                    pred_pct = int(pred.get("conclusao_pct", "0") or "0")
                    if edit_pct > pred_pct:
                        self.cron_error = (
                            f"⚠ Avanço limitado: predecessora está em {pred_pct}%. "
                            f"Não é possível salvar com {edit_pct}%."
                        )
                        self.cron_saving = False
                        return
            edit_dependencia = self.cron_edit_dependencia.strip()
            # Clear dependencia_id when type is sem_dep
            edit_dependencia_id = (
                (self.cron_edit_dependencia_id.strip() or None)
                if edit_dep_tipo != "sem_dep" else None
            )
            edit_observacoes = self.cron_edit_observacoes.strip()
            edit_nivel = self.cron_edit_nivel or "macro"
            edit_parent_id = self.cron_edit_parent_id or None
            edit_peso = max(1, int(self.cron_edit_peso or 100))
            try:
                edit_total_qty = float(self.cron_edit_total_qty.replace(",", ".")) if self.cron_edit_total_qty.strip() else 0.0
            except Exception:
                edit_total_qty = 0.0
            edit_unidade = self.cron_edit_unidade.strip()
            try:
                edit_dias_planejados = int(self.cron_edit_dias_planejados.strip()) if self.cron_edit_dias_planejados.strip() else 0
            except Exception:
                edit_dias_planejados = 0
            edit_status_atividade = self.cron_edit_status_atividade or "nao_iniciada"
            edit_tipo_medicao = self.cron_edit_tipo_medicao or "quantidade"
            try:
                edit_efetivo_alocado = int(self.cron_edit_efetivo_alocado.strip()) if self.cron_edit_efetivo_alocado.strip() else 0
            except Exception:
                edit_efetivo_alocado = 0

            # ── Validação de consistência de datas ────────────────────────────
            if edit_inicio and edit_termino:
                try:
                    from datetime import date as _date
                    d_inicio = _date.fromisoformat(edit_inicio)
                    d_termino = _date.fromisoformat(edit_termino)
                    if d_termino < d_inicio:
                        self.cron_error = "⚠ A data de término não pode ser anterior ao início."
                        self.cron_saving = False
                        return
                except Exception:
                    pass

            if edit_parent_id and edit_inicio:
                parent_row = next((r for r in self.cron_rows if r["id"] == edit_parent_id), None)
                if parent_row:
                    parent_inicio = parent_row.get("inicio_iso", "")
                    warnings: list = []
                    if parent_inicio:
                        try:
                            from datetime import date as _date
                            d_child = _date.fromisoformat(edit_inicio)
                            d_parent = _date.fromisoformat(parent_inicio)
                            if d_child < d_parent:
                                warnings.append(
                                    f"⚠ A data de início ({edit_inicio}) é anterior ao início da atividade pai "
                                    f"({parent_inicio})."
                                )
                        except Exception:
                            pass
                    if warnings:
                        self.cron_error = " | ".join(warnings)
                        self.cron_saving = False
                        return

            # Capture old values for full diff log
            old_snapshot: Dict[str, str] = {}
            if edit_id:
                old_row = next((r for r in self.cron_rows if r["id"] == edit_id), {})
                old_inicio = old_row.get("inicio_iso", "")
                old_termino = old_row.get("termino_iso", "")
                old_snapshot = {
                    "inicio_previsto":  old_inicio,
                    "termino_previsto": old_termino,
                    "conclusao_pct":    old_row.get("conclusao_pct", ""),
                    "responsavel":      old_row.get("responsavel", ""),
                    "peso_pct":         old_row.get("peso_pct", ""),
                    "critico":          old_row.get("critico", ""),
                    "nivel":            old_row.get("nivel", ""),
                    "fase_macro":       old_row.get("fase_macro", ""),
                    "fase":             old_row.get("fase", ""),
                    "observacoes":      old_row.get("observacoes", ""),
                    "total_qty":        old_row.get("total_qty", ""),
                    "unidade":          old_row.get("unidade", ""),
                    "dias_planejados":  old_row.get("dias_planejados", ""),
                    "status_atividade": old_row.get("status_atividade", ""),
                    "tipo_medicao":     old_row.get("tipo_medicao", ""),
                }

        # Sempre busca GlobalState para client_id + fallback de contrato (fora do lock)
        try:
            gs = await self.get_state(GlobalState)
            async with self:
                if not contrato:
                    contrato = str(gs.selected_contrato or gs.selected_project or "")
                client_id = str(gs.current_client_id or "")
                username = str(gs.current_user_name or "")
        except Exception as e:
            logger.error(f"save_cron_activity get_state error: {e}", exc_info=True)
            async with self:
                self.cron_saving = False
                self.cron_error = "Erro ao obter estado. Tente novamente."
            return

        try:
            data: Dict[str, Any] = {
                "contrato":          contrato,
                "fase_macro":        edit_fase_macro,
                "fase":              edit_fase,
                "atividade":         atividade_nome,
                "responsavel":       edit_responsavel,
                "inicio_previsto":   edit_inicio,
                "termino_previsto":  edit_termino,
                "conclusao_pct":     edit_pct,
                "critico":           edit_critico,
                "dep_tipo":          edit_dep_tipo,
                "dependencia":       edit_dependencia,
                "dependencia_id":    edit_dependencia_id,
                "observacoes":       edit_observacoes,
                "nivel":             edit_nivel,
                "parent_id":         edit_parent_id,
                "peso_pct":          edit_peso,
                "total_qty":         edit_total_qty,
                "unidade":           edit_unidade,
                "dias_planejados":   edit_dias_planejados,
                "status_atividade":  edit_status_atividade,
                "tipo_medicao":      edit_tipo_medicao,
                "efetivo_alocado":   edit_efetivo_alocado,
                "client_id":         client_id,
            }

            if edit_id:
                if pending_review_id:
                    data["pendente_aprovacao"] = False
                sb_update("hub_atividades", filters={"id": edit_id}, data=data)
                action = f"Atividade '{atividade_nome}' {'aprovada' if pending_review_id else 'atualizada'}"

                # ── Cascade renome de macro → fase_macro em todos os filhos ──────
                # Quando uma macro é renomeada, todas as micros e subs filhas precisam
                # ter fase_macro atualizado para manter a hierarquia consistente.
                # Sem isso, a micro exibe "Montagem X" mas a macro chama "Estrutura Y".
                if edit_nivel == "macro" and not pending_review_id:
                    old_fase_macro = old_snapshot.get("fase_macro", "")
                    if old_fase_macro != edit_fase_macro:
                        try:
                            micros = sb_select("hub_atividades", filters={"parent_id": edit_id}, limit=200)
                            for micro in micros:
                                mid = micro.get("id", "")
                                if mid:
                                    sb_update("hub_atividades", filters={"id": mid}, data={"fase_macro": edit_fase_macro})
                                    # Cascade subs (sub → micro → macro)
                                    subs = sb_select("hub_atividades", filters={"parent_id": mid}, limit=200)
                                    for sub in subs:
                                        sid = sub.get("id", "")
                                        if sid:
                                            sb_update("hub_atividades", filters={"id": sid}, data={"fase_macro": edit_fase_macro})
                            logger.info(f"✅ Cascade fase_macro: '{old_fase_macro}' → '{edit_fase_macro}' para filhos de {edit_id}")
                        except Exception as _casc_err:
                            logger.warning(f"⚠️ Cascade fase_macro error (non-fatal): {_casc_err}")

                # ── Full diff log → hub_cronograma_log + hub_timeline ─────────
                if not pending_review_id:
                    new_snapshot = {
                        "inicio_previsto":  edit_inicio or "",
                        "termino_previsto": edit_termino or "",
                        "conclusao_pct":    str(edit_pct),
                        "responsavel":      edit_responsavel,
                        "peso_pct":         str(edit_peso),
                        "critico":          "1" if edit_critico else "0",
                        "nivel":            edit_nivel,
                        "fase_macro":       edit_fase_macro,
                        "fase":             edit_fase,
                        "observacoes":      edit_observacoes,
                        "total_qty":        str(edit_total_qty),
                        "unidade":          edit_unidade,
                        "dias_planejados":  str(edit_dias_planejados),
                        "status_atividade": edit_status_atividade,
                        "tipo_medicao":     edit_tipo_medicao,
                    }
                    _log_schedule_diff_async(
                        contrato=contrato,
                        atividade_id=edit_id,
                        atividade_nome=atividade_nome,
                        old_row=old_snapshot,
                        new_row=new_snapshot,
                        autor=username,
                        client_id=client_id,
                    )
            else:
                sb_insert("hub_atividades", data)
                action = f"Atividade '{atividade_nome}' criada"

            # ── Auto-derive macro dates from micros ───────────────────────────
            if edit_nivel == "micro" and edit_parent_id:
                _recalc_macro_dates(edit_parent_id, contrato, client_id)

            audit_log(
                category=AuditCategory.DATA_EDIT,
                action=action,
                username=username,
                entity_type="hub_atividades",
                entity_id=edit_id or "new",
                metadata={"contrato": contrato, "atividade": atividade_nome},
            )

        except Exception as e:
            logger.error(f"save_cron_activity error: {e}")
            async with self:
                self.cron_error = f"Erro: {str(e)[:120]}"
                self.cron_saving = False
            return

        # Reload
        async with self:
            self.cron_show_dialog = False
            self.cron_saving = False
            self.cron_pending_review_id = ""

        yield HubState.load_cronograma(contrato)

    # ── Delete ────────────────────────────────────────────────────────────────

    def request_cron_delete(self, row_id: str):
        row = next((r for r in self.cron_rows if r["id"] == row_id), None)
        self.cron_delete_id = row_id
        self.cron_delete_name = row.get("atividade", row_id) if row else row_id
        self.cron_show_delete = True

    def cancel_cron_delete(self):
        self.cron_delete_id = ""
        self.cron_show_delete = False

    @rx.event(background=True)
    async def confirm_cron_delete(self):
        row_id = ""
        name = ""
        contrato = ""
        backup_rows: list = []
        async with self:
            row_id = str(self.cron_delete_id)
            name = str(self.cron_delete_name)
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
            self.cron_show_delete = False
            self.cron_delete_id = ""
            # Optimistic UI: backup + remove imediatamente antes do DB call
            backup_rows = list(self.cron_rows)
            self.cron_rows = [r for r in self.cron_rows if r.get("id") != row_id]

        if not row_id:
            logger.error("confirm_cron_delete: row_id vazio, abortando.")
            yield rx.toast.error("ID da atividade não encontrado.", duration=4000)
            return

        delete_ok = False
        try:
            logger.info(f"confirm_cron_delete: deleting id={row_id!r} name={name!r}")
            # 1. Cascade delete: children (micros) and grandchildren (subs) of the target
            def _delete_activity_and_children(act_id: str):
                """Recursively delete an activity and all its descendants."""
                try:
                    grandchildren = sb_select("hub_atividades", filters={"parent_id": act_id}, limit=200) or []
                    for gc in grandchildren:
                        gc_id = str(gc.get("id", ""))
                        if gc_id:
                            _delete_activity_and_children(gc_id)
                except Exception:
                    pass
                try:
                    sb_delete("hub_atividade_historico", filters={"atividade_id": act_id})
                except Exception:
                    pass
                try:
                    sb_delete("hub_cronograma_log", filters={"atividade_id": act_id})
                except Exception:
                    pass
                sb_delete("hub_atividades", filters={"id": act_id})

            try:
                children = sb_select("hub_atividades", filters={"parent_id": row_id}, limit=200) or []
                for child in children:
                    child_id = str(child.get("id", ""))
                    if child_id:
                        _delete_activity_and_children(child_id)
            except Exception as child_ex:
                logger.warning(f"confirm_cron_delete: erro ao deletar filhos de {row_id}: {child_ex}")
            # 2. Delete deps of the target row itself
            try:
                sb_delete("hub_atividade_historico", filters={"atividade_id": row_id})
            except Exception:
                pass
            try:
                sb_delete("hub_cronograma_log", filters={"atividade_id": row_id})
            except Exception:
                pass
            # 3. Delete the row
            delete_ok = sb_delete("hub_atividades", filters={"id": row_id})
            if delete_ok:
                audit_log(
                    category=AuditCategory.DATA_DELETE,
                    action=f"Atividade '{name}' excluída",
                    username="",
                    entity_type="hub_atividades",
                    entity_id=row_id,
                )
            else:
                logger.error(f"confirm_cron_delete: sb_delete retornou False para id={row_id}")
                async with self:
                    self.cron_rows = backup_rows  # rollback optimistic remove
                yield rx.toast.error("Erro ao excluir atividade no banco de dados.", duration=4000)
        except Exception as e:
            logger.error(f"confirm_cron_delete error: {e}")
            async with self:
                self.cron_rows = backup_rows  # rollback optimistic remove
            yield rx.toast.error(f"Erro ao excluir: {str(e)[:100]}", duration=4000)

        if contrato:
            yield HubState.load_cronograma(contrato)

    # ── Approve / Reject pending activities ───────────────────────────────────

    @rx.event(background=True)
    async def approve_pending_activity(self, activity_id: str):
        """Approve a pending activity (set pendente_aprovacao=False)."""
        contrato = ""
        async with self:
            self.cron_approve_loading = True
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
        try:
            sb_update("hub_atividades", filters={"id": activity_id}, data={"pendente_aprovacao": False})
        except Exception as e:
            logger.error(f"approve_pending_activity error: {e}")
        async with self:
            self.cron_approve_loading = False
        if contrato:
            yield HubState.load_cronograma(contrato)

    @rx.event(background=True)
    async def reject_pending_activity(self, activity_id: str):
        """Reject (delete) a pending activity."""
        contrato = ""
        async with self:
            self.cron_approve_loading = True
            contrato = self.cron_rows[0].get("contrato", "") if self.cron_rows else ""
        try:
            sb_delete("hub_atividades", filters={"id": activity_id})
        except Exception as e:
            logger.error(f"reject_pending_activity error: {e}")
        async with self:
            self.cron_approve_loading = False
        if contrato:
            yield HubState.load_cronograma(contrato)

    # ── IA Climate Analysis ────────────────────────────────────────────────────

    @rx.event(background=True)
    async def analyze_climate_impact(self):
        """Cross-reference scheduled activities with weather forecast via IA."""
        from bomtempo.state.global_state import GlobalState

        weather = {}
        rows = []
        async with self:
            gs = await self.get_state(GlobalState)
            self.cron_climate_loading = True
            self.cron_climate_analysis = ""
            weather = dict(gs.weather_data) if gs.weather_data else {}
            rows = list(self.cron_rows)

        # Build context
        from datetime import date
        today_iso = date.today().isoformat()
        weather_summary = ""
        if weather and weather.get("daily_time"):
            lines = []
            for i, dt in enumerate(weather["daily_time"][:7]):
                prob = weather["daily_rain_prob"][i] if i < len(weather.get("daily_rain_prob", [])) else 0
                rain = weather["daily_rain_sum"][i] if i < len(weather.get("daily_rain_sum", [])) else 0
                lines.append(f"  {dt}: chuva {rain:.1f}mm, prob. {prob}%")
            weather_summary = "\n".join(lines)
        else:
            weather_summary = "  Dados climáticos não disponíveis"

        activities_summary = ""
        if rows:
            act_lines = []
            for r in rows[:20]:
                act_lines.append(
                    f"  • {r['atividade']} ({r['fase_macro']}) | {r['inicio_previsto']} → {r['termino_previsto']} | {r['conclusao_pct']}% | crítico: {'sim' if r['critico']=='1' else 'não'}"
                )
            activities_summary = "\n".join(act_lines)
        else:
            activities_summary = "  Nenhuma atividade cadastrada"

        messages = [
            {
                "role": "user",
                "content": (
                    f"Você é um engenheiro de obras sênior analisando impacto climático em um cronograma de construção.\n\n"
                    f"DATA ATUAL: {today_iso}\n\n"
                    f"PREVISÃO CLIMÁTICA (próximos 7 dias):\n{weather_summary}\n\n"
                    f"ATIVIDADES DO CRONOGRAMA:\n{activities_summary}\n\n"
                    f"Analise quais atividades previstas para os próximos 7 dias serão impactadas pela chuva, "
                    f"especialmente as críticas. Dê recomendações práticas: o que antecipar, o que proteger, "
                    f"o que reagendar. Seja direto e objetivo, máximo 200 palavras."
                ),
            }
        ]

        try:
            import queue as _queue
            import threading
            result_queue: _queue.Queue = _queue.Queue()

            def _run_analysis():
                try:
                    # Use direct AI call
                    from bomtempo.core.ai_client import ai_client
                    full_text = ""
                    for chunk in ai_client.query_stream(messages):
                        full_text += chunk
                    result_queue.put(("ok", full_text))
                except Exception as ex:
                    result_queue.put(("err", str(ex)))

            t = threading.Thread(target=_run_analysis, daemon=True)
            t.start()
            t.join(timeout=60)

            if not result_queue.empty():
                status, text = result_queue.get()
                async with self:
                    self.cron_climate_analysis = text if status == "ok" else f"Erro na análise: {text[:200]}"
                if status == "ok":
                    yield rx.toast.success("Análise climática concluída — veja o resultado abaixo.", duration=5000)
                    yield rx.call_script("document.getElementById('climate-analysis-panel')?.scrollIntoView({behavior:'smooth',block:'start'})")
            else:
                async with self:
                    self.cron_climate_analysis = "Tempo esgotado ao consultar IA. Tente novamente."
                yield rx.toast.warning("Análise climática: tempo esgotado.", duration=4000)

        except Exception as e:
            logger.error(f"analyze_climate_impact error: {e}")
            async with self:
                self.cron_climate_analysis = f"Erro: {str(e)[:200]}"
        finally:
            async with self:
                self.cron_climate_loading = False

    def clear_climate_analysis(self):
        self.cron_climate_analysis = ""

    # ── IA Import from Excel / PDF ─────────────────────────────────────────────

    async def import_cronograma_ia(self, files: list[rx.UploadFile]):
        """Receive uploaded file via rx.upload, parse content, send to IA, propose activities."""
        from bomtempo.state.global_state import GlobalState
        import json

        # Regular async upload handler — state writes are direct (no async with self needed)
        self.cron_import_loading = True
        self.cron_import_error = ""
        self.cron_import_preview = []
        # Use cron_contrato (set by load_cronograma) as the authoritative source
        contrato = self.cron_contrato or (self.cron_rows[0].get("contrato", "") if self.cron_rows else "")
        yield  # flush loading=True to UI

        file_text = ""
        if files:
            upload_file = files[0]
            name = upload_file.filename
            raw_bytes = await upload_file.read()
            try:
                ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
                if ext in ("xlsx", "xls"):
                    import io
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
                    all_lines = []
                    # Itera TODAS as abas do Excel
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        all_lines.append(f"\n=== ABA: {sheet_name} ===")
                        for row in ws.iter_rows(values_only=True):
                            cells = [str(c) if c is not None else "" for c in row]
                            row_text = " | ".join(cells).strip(" |")
                            if row_text:  # ignora linhas completamente vazias
                                all_lines.append(row_text)
                    file_text = "\n".join(all_lines[:400])
                elif ext == "csv":
                    file_text = raw_bytes.decode("utf-8", errors="replace")[:8000]
                else:
                    # PDF or other: decode as text best-effort
                    file_text = raw_bytes.decode("utf-8", errors="replace")[:8000]
            except Exception as ex:
                self.cron_import_error = f"Erro ao ler arquivo: {ex}"
                self.cron_import_loading = False
                return

        if not file_text.strip():
            self.cron_import_error = "Arquivo vazio ou não legível."
            self.cron_import_loading = False
            return

        prompt = f"""Você é um assistente especialista em gestão de obras e cronogramas (engenharia civil, elétrica, solar, instalações). Analise o arquivo abaixo e extraia EXATAMENTE as atividades explicitamente listadas.

═══════════════════════════════════════════════════
ESTRUTURA DE HIERARQUIA (CRÍTICO — leia com atenção)
═══════════════════════════════════════════════════
O sistema organiza atividades em dois níveis:
  • nivel="macro" → disciplina/fase principal (ex: "1. Fundações", "2. Estrutura", "Elétrica CA")
  • nivel="micro" → sub-atividade dentro de uma macro (ex: "1.1 Escavação", "1.2 Armação")

REGRA OBRIGATÓRIA: cada atividade micro DEVE ter o mesmo "fase_macro" que sua macro pai.
  Exemplo correto:
    {{ "atividade": "Elétrica CA", "nivel": "macro", "fase_macro": "Elétrica CA", "fase": "2" }}
    {{ "atividade": "Do QGCA para o QGBT", "nivel": "micro", "fase_macro": "Elétrica CA", "fase": "2.1" }}
    {{ "atividade": "Cabeamento", "nivel": "micro", "fase_macro": "Elétrica CA", "fase": "2.2" }}
  Exemplo ERRADO (não faça):
    {{ "atividade": "Do QGCA para o QGBT", "nivel": "micro", "fase_macro": "Fundações" }}  ← fase_macro errada!

CAMPO "fase" = índice/número hierárquico da atividade no cronograma:
  - Macros: "1", "2", "3"... ou o nome da seção se não houver número
  - Micros: "1.1", "1.2", "2.3"... Se o arquivo não tiver numeração, infira pela posição
  - Nunca deixe "fase" vazio — use pelo menos "1" para a primeira macro, "1.1" para a primeira micro dela

═══════════════════════════════════════════════════
REGRAS GERAIS
═══════════════════════════════════════════════════
1. Extraia APENAS atividades EXPLICITAMENTE no arquivo. Se não encontrar, retorne lista vazia.
2. NÃO invente, NÃO duplique, NÃO crie atividades inferidas.
3. NÃO duplique: se a mesma atividade aparece em abas diferentes, use a com mais informações.

═══════════════════════════════════════════════════
DATAS E DURAÇÃO
═══════════════════════════════════════════════════
4. DATAS — extraia com atenção:
   a. Colunas "início", "start", "data início" → inicio_previsto
   b. Colunas "término", "fim", "end", "conclusão" → termino_previsto
   c. DD/MM/AAAA ou DD/MM/AA → YYYY-MM-DD (ex: 15/03/2025 → 2025-03-15)
   d. MM/DD/YYYY → YYYY-MM-DD
   e. "março 2025", "mar/25" → primeiro dia do mês (2025-03-01)
   f. GANTT visual (barras em colunas): leia os cabeçalhos das colunas onde a barra começa e termina
      - Seja assertivo: se o cabeçalho da primeira coluna com barra é "JAN/25" e última é "MAR/25",
        extraia inicio_previsto="2025-01-01" e termino_previsto="2025-03-31"
      - SOMENTE omita a data se não conseguir determinar a coluna com confiança
   g. Sem info de data → deixe vazio

5. dias_planejados:
   a. USE coluna explícita de duração ("dias", "duration", "prazo", "Dur.") se existir
   b. SE há inicio_previsto E termino_previsto → calcule como dias corridos (termino - inicio)
   c. SE Gantt: conte as colunas ocupadas pela barra × período da coluna (ex: 3 colunas mensais = 90 dias)
   d. Se não houver NENHUMA informação confiável → deixe 0

6. total_qty e unidade: extraia APENAS se houver coluna explícita ("qtd", "quantidade", "m²", etc.). Senão 0 e vazio.

═══════════════════════════════════════════════════
OUTROS CAMPOS
═══════════════════════════════════════════════════
7. Crítico: atividades marcadas em vermelho, negrito, asterisco ou flag → critico=true
8. responsavel: coluna de responsável/equipe/sub-empreiteiro, se existir
9. observacoes: qualquer nota relevante que não caiba nos outros campos

10. Confidence (0.0 a 1.0):
    - 1.0: arquivo estruturado, hierarquia clara, datas explícitas
    - 0.75-0.95: boa estrutura, pequenas ambiguidades
    - 0.5-0.74: hierarquia parcialmente clara, datas ausentes em muitos itens
    - <0.5: texto livre, Gantt difícil, estrutura muito ambígua

ARQUIVO:
{file_text[:7000]}

Retorne SOMENTE JSON válido, sem texto antes/depois, sem markdown:
{{
  "confidence": 0.9,
  "activities": [
    {{
      "atividade": "Nome exato da atividade conforme o arquivo",
      "fase_macro": "Nome da disciplina/fase principal — DEVE ser idêntico entre macro e suas micros",
      "fase": "Índice numérico ex: '1', '1.1', '2.3' — nunca vazio",
      "responsavel": "Responsável/equipe se disponível, senão vazio",
      "inicio_previsto": "YYYY-MM-DD ou vazio",
      "termino_previsto": "YYYY-MM-DD ou vazio",
      "dias_planejados": número inteiro ou 0,
      "total_qty": número ou 0,
      "unidade": "m, m², m³, kg, und, kWh, kW, kWp ou vazio",
      "critico": true ou false,
      "nivel": "macro" ou "micro",
      "observacoes": "notas relevantes ou vazio"
    }}
  ]
}}"""

        import asyncio, queue as _queue, threading
        result_q: _queue.Queue = _queue.Queue()

        def _call_ai():
            try:
                from bomtempo.core.ai_client import ai_client
                resp = ai_client.query([{"role": "user", "content": prompt}])
                result_q.put(("ok", resp))
            except Exception as ex:
                result_q.put(("err", str(ex)))

        loop = asyncio.get_running_loop()
        await loop.run_in_executor(get_ai_executor(), _call_ai)

        if result_q.empty():
            self.cron_import_error = "IA não respondeu a tempo. Tente novamente."
            self.cron_import_loading = False
            return

        status, raw = result_q.get()
        if status == "err":
            self.cron_import_error = f"Erro IA: {raw[:200]}"
            self.cron_import_loading = False
            return

        # Parse JSON from response
        try:
            # Strip markdown code fences if present
            cleaned = raw.strip()
            if cleaned.startswith("```"):
                cleaned = cleaned.split("\n", 1)[-1].rsplit("```", 1)[0]
            parsed = json.loads(cleaned)
            # Support both new format {"confidence": X, "activities": [...]} and legacy plain list
            if isinstance(parsed, dict):
                confidence = float(parsed.get("confidence", 0.5))
                proposals = parsed.get("activities", [])
            elif isinstance(parsed, list):
                confidence = 0.5  # legacy: assume medium confidence
                proposals = parsed
            else:
                raise ValueError("Resposta não é um objeto nem lista")
            if not isinstance(proposals, list):
                raise ValueError("Campo 'activities' não é uma lista")
        except Exception as ex:
            self.cron_import_error = f"IA retornou formato inválido: {ex}"
            self.cron_import_loading = False
            return

        # Dedup by normalized activity name to prevent duplicates from multi-sheet parsing
        seen_names: dict = {}
        deduped = []
        for p in proposals:
            name_key = str(p.get("atividade", "")).strip().lower()
            if not name_key:
                continue
            if name_key not in seen_names:
                seen_names[name_key] = p
                deduped.append(p)
            else:
                # Keep whichever has more data (prefer the one with dates)
                existing = seen_names[name_key]
                existing_has_dates = bool(existing.get("inicio_previsto") or existing.get("termino_previsto"))
                new_has_dates = bool(p.get("inicio_previsto") or p.get("termino_previsto"))
                if new_has_dates and not existing_has_dates:
                    # Replace with the version that has dates
                    seen_names[name_key] = p
                    deduped = [p if x is existing else x for x in deduped]

        # Confidence label for display
        if confidence >= 0.8:
            conf_label = f"Alta ({int(confidence * 100)}%)"
        elif confidence >= 0.5:
            conf_label = f"Média ({int(confidence * 100)}%)"
        else:
            conf_label = f"Baixa ({int(confidence * 100)}%) — revise com atenção"

        # Normalize proposals into display rows with temp IDs
        import uuid
        preview = []
        for i, p in enumerate(deduped[:50]):
            preview.append({
                "_tmp_id":          str(uuid.uuid4()),
                "atividade":        str(p.get("atividade", f"Atividade {i+1}")),
                "fase_macro":       str(p.get("fase_macro", "")),
                "fase":             str(p.get("fase", "")),
                "responsavel":      str(p.get("responsavel", "")),
                "inicio_previsto":  str(p.get("inicio_previsto", "")),
                "termino_previsto": str(p.get("termino_previsto", "")),
                "dias_planejados":  str(p.get("dias_planejados", 0) or 0),
                "total_qty":        str(p.get("total_qty", 0) or 0),
                "unidade":          str(p.get("unidade", "") or ""),
                "critico":          "1" if p.get("critico") else "0",
                "nivel":            str(p.get("nivel", "macro")),
                "observacoes":      str(p.get("observacoes", "")),
                "contrato":         contrato,
            })

        # Count items with 0 dias_planejados for consistency warning
        zero_days = sum(1 for r in preview if r["dias_planejados"] == "0")
        has_dates = sum(1 for r in preview if r["inicio_previsto"] or r["termino_previsto"])

        msg = f"{len(preview)} atividades identificadas. Confiança: {conf_label}."
        if zero_days > 0 and has_dates == 0:
            msg += f" Atenção: {zero_days} atividade(s) sem datas — preencha manualmente após importar."
        elif zero_days > 0:
            msg += f" {zero_days} atividade(s) sem duração calculada."

        self.cron_import_preview = preview
        self.cron_import_selected = [r["_tmp_id"] for r in preview]
        self.cron_import_confidence_label = conf_label
        self.cron_import_show = True
        self.cron_import_loading = False
        yield rx.toast.success(msg, duration=8000)

    def toggle_import_activity(self, tmp_id: str):
        if tmp_id in self.cron_import_selected:
            self.cron_import_selected = [x for x in self.cron_import_selected if x != tmp_id]
        else:
            new = list(self.cron_import_selected)
            new.append(tmp_id)
            self.cron_import_selected = new

    def select_all_import(self):
        self.cron_import_selected = [r["_tmp_id"] for r in self.cron_import_preview]

    def deselect_all_import(self):
        self.cron_import_selected = []

    def close_import_preview(self):
        self.cron_import_show = False
        self.cron_import_preview = []
        self.cron_import_selected = []
        self.cron_import_confidence_label = ""

    @rx.event(background=True)
    async def confirm_import_cronograma(self):
        """Bulk-insert selected proposals into hub_atividades."""
        from bomtempo.state.global_state import GlobalState

        client_id = ""
        selected_ids = set()
        to_insert = []
        contrato = ""

        working_days_str = "seg,ter,qua,qui,sex"
        async with self:
            gs = await self.get_state(GlobalState)
            client_id = str(gs.current_client_id or "")
            selected_ids = set(self.cron_import_selected)
            all_preview = list(self.cron_import_preview)
            # Fallback: se selected_ids estiver vazio mas há preview, assume todas selecionadas
            # (pode acontecer quando o botão "Selecionar todas" é clicado logo antes de confirmar)
            if not selected_ids and all_preview:
                selected_ids = {r["_tmp_id"] for r in all_preview}
            to_insert = [r for r in all_preview if r["_tmp_id"] in selected_ids]
            contrato = to_insert[0]["contrato"] if to_insert else ""
            working_days_str = self.cron_working_days_str
            self.cron_import_loading = True

        if not to_insert:
            async with self:
                self.cron_import_loading = False
            yield rx.toast.warning("Nenhuma atividade selecionada.", duration=3000)
            return

        logger.info(f"confirm_import_cronograma: inserindo {len(to_insert)} atividades para contrato={contrato!r}")

        errors = 0
        inserted = 0

        def _build_row(p: dict, parent_id: str = "") -> dict:
            dias = int(p.get("dias_planejados", 0) or 0)
            inicio = p.get("inicio_previsto", "") or None
            termino = p.get("termino_previsto", "") or None
            if inicio and dias and not termino:
                wd = _parse_dias_uteis(working_days_str)
                termino = _add_working_days(inicio, dias, wd)
            total_qty = 0.0
            try:
                total_qty = float(p.get("total_qty", 0) or 0)
            except Exception:
                pass
            unidade = p.get("unidade", "") or ""
            tipo_medicao = "quantidade" if total_qty > 0 else "percentual"
            row = {
                "contrato":          contrato,
                "fase_macro":        p.get("fase_macro", ""),
                "fase":              p.get("fase", ""),
                "atividade":         p.get("atividade", ""),
                "responsavel":       p.get("responsavel", "") or None,
                "inicio_previsto":   inicio,
                "termino_previsto":  termino,
                "conclusao_pct":     0,
                "critico":           p.get("critico", "0") == "1",
                "nivel":             p.get("nivel", "macro"),
                "peso_pct":          100,
                "dias_planejados":   dias,
                "total_qty":         total_qty,
                "unidade":           unidade,
                "observacoes":       p.get("observacoes", ""),
                "status_atividade":  "nao_iniciada",
                "tipo_medicao":      tipo_medicao,
                "client_id":         client_id or None,
            }
            if parent_id:
                row["parent_id"] = parent_id
            return row

        # Separate macros from micros; sort each group by "fase" index numerically
        def _fase_sort_key(p: dict) -> tuple:
            """Sort by hierarchical index: '1' < '1.1' < '1.2' < '2' < '2.1'"""
            fase = str(p.get("fase", "") or "")
            parts = []
            for seg in fase.split("."):
                try:
                    parts.append(int(seg))
                except ValueError:
                    parts.append(0)
            return tuple(parts) if parts else (9999,)

        macros = sorted(
            [p for p in to_insert if p.get("nivel", "macro") != "micro"],
            key=_fase_sort_key,
        )
        micros = sorted(
            [p for p in to_insert if p.get("nivel", "macro") == "micro"],
            key=_fase_sort_key,
        )

        # Insert macros first, build fase_macro → db_id map for parent linking
        fase_macro_to_id: dict = {}
        for p in macros:
            try:
                result = sb_insert("hub_atividades", _build_row(p))
                if result:
                    db_id = str(result.get("id", ""))
                    fm = p.get("fase_macro", "")
                    if db_id and fm and fm not in fase_macro_to_id:
                        fase_macro_to_id[fm] = db_id
                inserted += 1
            except Exception as ex:
                logger.warning(f"confirm_import macro error '{p.get('atividade','?')}': {ex}")
                errors += 1

        # Insert micros with parent_id resolved from fase_macro
        for p in micros:
            try:
                parent_id = fase_macro_to_id.get(p.get("fase_macro", ""), "")
                result = sb_insert("hub_atividades", _build_row(p, parent_id=parent_id))
                inserted += 1
            except Exception as ex:
                logger.warning(f"confirm_import micro error '{p.get('atividade','?')}': {ex}")
                errors += 1

        async with self:
            self.cron_import_show = False
            self.cron_import_preview = []
            self.cron_import_selected = []
            self.cron_import_loading = False

        if inserted:
            yield rx.toast.success(f"{inserted} atividade(s) importadas com sucesso!", duration=5000)
        if errors:
            yield rx.toast.warning(f"{errors} atividade(s) falharam na importação.", duration=5000)
        if contrato:
            yield HubState.load_cronograma(contrato)

    # ══════════════════════════════════════════════════════════════════════════
    # AUDITORIA — Load, gallery, lightbox
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def load_auditoria(self, contrato: str):
        async with self:
            self.audit_loading = True
            self.audit_images = []
            self.audit_open_category = ""

        # Captura client_id para isolamento de tenant
        client_id = ""
        try:
            from bomtempo.state.global_state import GlobalState
            _gs = await self.get_state(GlobalState)
            client_id = str(_gs.current_client_id or "")
        except Exception:
            pass

        try:
            # 1. Manual uploads in hub_auditoria_imgs
            _audit_filters: dict = {"contrato": contrato}
            if client_id:
                _audit_filters["client_id"] = client_id
            rows = sb_select(
                "hub_auditoria_imgs",
                filters=_audit_filters,
                order="created_at.desc",
                limit=500,
            )
            imgs = [
                {
                    "id":           _norm_str(r.get("id")),
                    "contrato":     _norm_str(r.get("contrato")),
                    "categoria":    _norm_str(r.get("categoria")),
                    "url":          _norm_str(r.get("url")),
                    "legenda":      _norm_str(r.get("legenda")),
                    "data_captura": _utc_date_to_br(_norm_str(r.get("data_captura"))),
                    "autor":        _norm_str(r.get("autor"), "—"),
                }
                for r in rows
            ]

            # 2. Integrate RDO photos from rdo_master (epi, ferramentas, evidencias)
            try:
                _rdo_filters: dict = {"contrato": contrato}
                if client_id:
                    _rdo_filters["client_id"] = client_id
                rdos = sb_select(
                    "rdo_master",
                    filters=_rdo_filters,
                    order="created_at.desc",
                    limit=200,
                )
                import json as _json
                for rdo in (rdos or []):
                    rdo_date = _utc_date_to_br(_norm_str(rdo.get("created_at", "") or rdo.get("data_rdo", "")))
                    rdo_autor = _norm_str(rdo.get("mestre_id", rdo.get("responsavel_tecnico", "RDO")))
                    rdo_id = _norm_str(rdo.get("id_rdo", rdo.get("id", "")))

                    # EPI photo
                    epi_url = _norm_str(rdo.get("epi_foto_url", ""))
                    if epi_url:
                        imgs.append({
                            "id":           f"rdo_epi_{rdo_id}",
                            "contrato":     contrato,
                            "categoria":    "equipe",
                            "url":          epi_url,
                            "legenda":      f"EPI — RDO {rdo_id[:8]}",
                            "data_captura": rdo_date,
                            "autor":        rdo_autor,
                        })

                    # Ferramentas photo
                    ferr_url = _norm_str(rdo.get("ferramentas_foto_url", ""))
                    if ferr_url:
                        imgs.append({
                            "id":           f"rdo_ferr_{rdo_id}",
                            "contrato":     contrato,
                            "categoria":    "ferramentas",
                            "url":          ferr_url,
                            "legenda":      f"Ferramentas — RDO {rdo_id[:8]}",
                            "data_captura": rdo_date,
                            "autor":        rdo_autor,
                        })

                    # Evidence photos (jsonb array)
                    evidencias_raw = rdo.get("evidencias") or []
                    if isinstance(evidencias_raw, str):
                        try:
                            evidencias_raw = _json.loads(evidencias_raw)
                        except Exception:
                            evidencias_raw = []
                    for idx, ev in enumerate(evidencias_raw or []):
                        ev_url = _norm_str(ev.get("foto_url", "") if isinstance(ev, dict) else "")
                        if ev_url:
                            legenda = _norm_str(ev.get("legenda", "") if isinstance(ev, dict) else "")
                            imgs.append({
                                "id":           f"rdo_ev_{rdo_id}_{idx}",
                                "contrato":     contrato,
                                "categoria":    "gerais",
                                "url":          ev_url,
                                "legenda":      legenda or f"Foto {idx+1} — RDO {rdo_id[:8]}",
                                "data_captura": rdo_date,
                                "autor":        rdo_autor,
                            })
            except Exception as e2:
                logger.warning(f"load_auditoria RDO integration error (non-fatal): {e2}")

        except Exception as e:
            logger.error(f"load_auditoria error: {e}")
            imgs = []

        async with self:
            self.audit_images = imgs
            self.audit_loading = False
            self._audit_loaded_contrato = contrato

    def open_audit_category(self, slug: str):
        # Toggle: clicking same category closes it
        self.audit_open_category = "" if self.audit_open_category == slug else slug

    def close_audit_category(self):
        self.audit_open_category = ""
        self.audit_lightbox_url = ""

    def open_lightbox(self, img_id: str):
        img = next((i for i in self.audit_images if i["id"] == img_id), None)
        if img:
            self.audit_lightbox_url = img["url"]
            self.audit_lightbox_legenda = img["legenda"]
            self.audit_lightbox_data = img["data_captura"]
            self.audit_lightbox_autor = img["autor"]

    def close_lightbox(self):
        self.audit_lightbox_url = ""

    def open_audit_upload(self, slug: str):
        self.audit_upload_category = slug
        self.audit_upload_url = ""
        self.audit_upload_legenda = ""
        self.audit_upload_error = ""
        self.audit_show_upload = True

    def close_audit_upload(self):
        self.audit_show_upload = False

    def set_audit_upload_url(self, v: str): self.audit_upload_url = v
    def set_audit_upload_legenda(self, v: str): self.audit_upload_legenda = v

    @rx.event(background=True)
    async def save_audit_image(self):
        from bomtempo.state.global_state import GlobalState

        contrato = ""
        autor = ""
        upload_category = ""
        upload_url = ""
        upload_legenda = ""

        async with self:
            if not self.audit_upload_url.strip():
                self.audit_upload_error = "URL da imagem é obrigatória."
                return
            self.audit_uploading = True
            self.audit_upload_error = ""
            contrato = self.audit_images[0]["contrato"] if self.audit_images else ""
            upload_category = str(self.audit_upload_category)
            upload_url = str(self.audit_upload_url).strip()
            upload_legenda = str(self.audit_upload_legenda).strip()

        try:
            gs = await self.get_state(GlobalState)
            autor = str(gs.current_user_name or "")
            async with self:
                if not contrato:
                    contrato = str(gs.selected_contrato or gs.selected_project or "")
        except Exception as e:
            logger.error(f"save_audit_image get_state error: {e}", exc_info=True)
            async with self:
                self.audit_uploading = False
                self.audit_upload_error = "Erro ao obter estado. Tente novamente."
            return

        from datetime import date as _date
        try:
            sb_insert("hub_auditoria_imgs", {
                "contrato":     contrato,
                "categoria":    upload_category,
                "url":          upload_url,
                "legenda":      upload_legenda,
                "autor":        autor,
                "data_captura": _date.today().isoformat(),
                "client_id":    str(gs.current_client_id or ""),
            })
        except Exception as e:
            logger.error(f"save_audit_image error: {e}")
            async with self:
                self.audit_upload_error = f"Erro: {str(e)[:100]}"
                self.audit_uploading = False
            return

        async with self:
            self.audit_show_upload = False
            self.audit_uploading = False

        yield HubState.load_auditoria(contrato)

    @rx.event(background=True)
    async def delete_audit_image(self, img_id: str):
        contrato = ""
        async with self:
            contrato = self.audit_images[0]["contrato"] if self.audit_images else ""
        try:
            sb_delete("hub_auditoria_imgs", filters={"id": img_id})
        except Exception as e:
            logger.error(f"delete_audit_image error: {e}")
        if contrato:
            yield HubState.load_auditoria(contrato)

    # ══════════════════════════════════════════════════════════════════════════
    # TIMELINE — Load & post
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def load_timeline(self, contrato: str):
        async with self:
            self.timeline_loading = True
            self.timeline_entries = []

        # Captura client_id para isolamento de tenant
        _tl_client_id = ""
        try:
            from bomtempo.state.global_state import GlobalState
            _tl_gs = await self.get_state(GlobalState)
            _tl_client_id = str(_tl_gs.current_client_id or "")
        except Exception:
            pass

        try:
            _tl_filters: dict = {"contrato": contrato}
            if _tl_client_id:
                _tl_filters["client_id"] = _tl_client_id
            rows = sb_select(
                "hub_timeline",
                filters=_tl_filters,
                order="created_at.desc",
                limit=200,
            )
            entries = [
                {
                    "id":              _norm_str(r.get("id")),
                    "contrato":        _norm_str(r.get("contrato")),
                    "tipo":            _norm_str(r.get("tipo"), "Atualização"),
                    "titulo":          _norm_str(r.get("titulo")),
                    "descricao":       _norm_str(r.get("descricao")),
                    "autor":           _norm_str(r.get("autor"), "—"),
                    "created_at":      _utc_to_brt(_norm_str(r.get("created_at"))),
                    "is_document":     "1" if r.get("is_document") else "0",
                    "is_cost":         "1" if r.get("is_cost") else "0",
                    "custo_valor":     _norm_str(r.get("custo_valor")),
                    "custo_categoria": _norm_str(r.get("custo_categoria")),
                    "anexo_url":       _norm_str(r.get("anexo_url")),
                    "anexo_nome":      _norm_str(r.get("anexo_nome")),
                }
                for r in rows
            ]
        except Exception as e:
            logger.error(f"load_timeline error: {e}")
            entries = []

        # Load mention users list — filtered by tenant
        try:
            _login_filters = {}
            if _tl_client_id:
                _login_filters["client_id"] = _tl_client_id
            login_rows = sb_select("login", filters=_login_filters if _login_filters else None, limit=100) or []
            mention_users = [str(r.get("user", "")).strip() for r in login_rows if r.get("user")]
        except Exception:
            mention_users = []

        async with self:
            self.timeline_entries = entries
            self.timeline_loading = False
            self._timeline_loaded_contrato = contrato
            if mention_users:
                self.tl_mention_users = mention_users

    def set_tl_entry_type(self, v: str): self.tl_entry_type = v
    def set_tl_titulo(self, v: str): self.tl_titulo = v
    def set_tl_descricao(self, v: str): self.tl_descricao = v
    def set_tl_filter_tipo(self, v: str):
        self.tl_filter_tipo = "" if self.tl_filter_tipo == v else v
    def set_tl_search(self, v: str): self.tl_search = v
    def set_tl_search_input(self, v: str): self.tl_search_input = v
    def commit_tl_search(self, _v: str = ""):
        self.tl_search = self.tl_search_input
    def handle_tl_search_key(self, key: str):
        if key == "Enter":
            self.tl_search = self.tl_search_input
    def set_tl_custo_valor(self, v: str): self.tl_custo_valor = v
    def set_tl_custo_categoria(self, v: str): self.tl_custo_categoria = v

    async def upload_tl_anexo(self, files: list[rx.UploadFile]):
        """Upload file attachment to Supabase Storage bucket 'timeline-anexos'."""
        if not files:
            return
        file = files[0]
        self.tl_uploading_anexo = True
        yield

        try:
            import asyncio as _asyncio
            import re as _re
            from datetime import datetime as _dt
            from bomtempo.core.supabase_client import sb_storage_ensure_bucket, sb_storage_upload

            data = await file.read()
            nome = getattr(file, "filename", getattr(file, "name", "arquivo"))
            safe_name = _re.sub(r"[^\w\.\-]", "_", nome)
            path = f"{_dt.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"

            loop = _asyncio.get_running_loop()
            await loop.run_in_executor(get_db_executor(), lambda: sb_storage_ensure_bucket("timeline-anexos", public=True))
            url = await loop.run_in_executor(get_db_executor(), lambda: sb_storage_upload("timeline-anexos", path, data, "application/octet-stream"))

            self.tl_anexo_url = url or ""
            self.tl_anexo_nome = nome
            self.tl_uploading_anexo = False
        except Exception as e:
            logger.error(f"upload_tl_anexo error: {e}")
            self.tl_uploading_anexo = False
            self.tl_error = f"Erro no upload: {str(e)[:80]}"

    @rx.event(background=True)
    async def submit_timeline_entry(self):
        from bomtempo.state.global_state import GlobalState

        contrato = ""
        autor = ""
        entry_tipo = ""
        entry_titulo = ""
        entry_descricao = ""
        entry_is_document = False
        entry_is_cost = False
        entry_custo_valor = ""
        entry_custo_categoria = ""
        entry_anexo_url = ""
        entry_anexo_nome = ""
        entry_mencoes: list = []

        async with self:
            if not self.tl_titulo.strip():
                self.tl_error = "Título é obrigatório."
                return
            self.tl_submitting = True
            self.tl_error = ""
            contrato = self.timeline_entries[0]["contrato"] if self.timeline_entries else ""
            entry_tipo = str(self.tl_entry_type)
            entry_titulo = str(self.tl_titulo).strip()
            entry_descricao = str(self.tl_descricao).strip()
            entry_is_document = entry_tipo == "Documento"
            entry_is_cost = entry_tipo == "Custo"
            entry_custo_valor = str(self.tl_custo_valor)
            entry_custo_categoria = str(self.tl_custo_categoria)
            entry_anexo_url = str(self.tl_anexo_url)
            entry_anexo_nome = str(self.tl_anexo_nome)
            # Extract @mentions from title + description
            import re as _re
            raw_text = f"{entry_titulo} {entry_descricao}"
            entry_mencoes = list(set(_re.findall(r"@(\w+)", raw_text)))

        try:
            gs = await self.get_state(GlobalState)
            autor = str(gs.current_user_name or "")
            async with self:
                if not contrato:
                    contrato = str(gs.selected_contrato or gs.selected_project or "")
        except Exception as e:
            logger.error(f"submit_timeline_entry get_state error: {e}", exc_info=True)
            async with self:
                self.tl_submitting = False
                self.tl_error = "Erro ao obter estado. Tente novamente."
            return

        try:
            tl_result = sb_insert("hub_timeline", {
                "contrato":        contrato,
                "tipo":            entry_tipo,
                "titulo":          entry_titulo,
                "descricao":       entry_descricao,
                "autor":           autor,
                "mencoes":         entry_mencoes,
                "is_document":     entry_is_document,
                "is_cost":         entry_is_cost,
                "custo_valor":     float(entry_custo_valor.replace(",", ".")) if entry_is_cost and entry_custo_valor else None,
                "custo_categoria": entry_custo_categoria if entry_is_cost else None,
                "anexo_url":       entry_anexo_url or None,
                "anexo_nome":      entry_anexo_nome or None,
                "client_id":       str(gs.current_client_id or ""),
            })
            audit_log(
                category=AuditCategory.DATA_EDIT,
                action=f"Timeline entry: [{entry_tipo}] {entry_titulo[:60]}",
                username=autor,
                entity_type="hub_timeline",
                entity_id=contrato,
            )
            # ── Document Intelligence: analisa documentos em background ──────
            if entry_is_document and entry_anexo_url:
                try:
                    from bomtempo.core.document_intel_service import DocumentIntelService
                    tl_id = (tl_result or {}).get("id", "") if tl_result else ""
                    DocumentIntelService.trigger_analysis(
                        timeline_id=tl_id,
                        contrato=contrato,
                        client_id=str(gs.current_client_id or ""),
                        titulo=entry_titulo,
                        descricao=entry_descricao,
                        anexo_url=entry_anexo_url,
                        anexo_nome=entry_anexo_nome,
                    )
                except Exception as _di_err:
                    logger.warning(f"[DocIntel] trigger_analysis falhou (não crítico): {_di_err}")
            # Integração financeira: custo na timeline → fin_custos (gasto não previsto)
            if entry_is_cost and entry_custo_valor:
                try:
                    from datetime import date as _tl_date
                    _custo_float = float(entry_custo_valor.replace(",", "."))
                    sb_insert("fin_custos", {
                        "contrato":       contrato,
                        "categoria_nome": entry_custo_categoria or "Outros",
                        "descricao":      f"[Timeline] {entry_titulo}",
                        "valor_previsto": 0.0,
                        "valor_executado": _custo_float,
                        "status":         "extrabudget",
                        "data":           _tl_date.today().isoformat(),
                        "criado_por":     autor,
                        "client_id":      str(gs.current_client_id or ""),
                    })
                    logger.info(f"💰 Custo timeline → fin_custos: R$ {_custo_float} [{entry_custo_categoria}] — {contrato}")
                except Exception as _fc_err:
                    logger.warning(f"⚠️ Falha ao propagar custo timeline para fin_custos: {_fc_err}")

            # Create notifications for @mentioned users
            if entry_mencoes:
                _notif_msg = f"@{autor} mencionou você: [{entry_tipo}] {entry_titulo[:80]}"
                for mentioned_user in entry_mencoes:
                    if mentioned_user:
                        try:
                            sb_insert("user_notifications", {
                                "recipient": mentioned_user,
                                "sender": autor,
                                "message": _notif_msg,
                                "source_type": "mention",
                                "source_id": contrato,
                                "contrato": contrato,
                                "read": False,
                                "client_id": str(gs.current_client_id or ""),
                            })
                        except Exception as ne:
                            logger.warning(f"Failed to create notification for @{mentioned_user}: {ne}")
        except Exception as e:
            logger.error(f"submit_timeline_entry error: {e}")
            async with self:
                self.tl_error = f"Erro: {str(e)[:100]}"
                self.tl_submitting = False
            return

        async with self:
            self.tl_titulo = ""
            self.tl_descricao = ""
            self.tl_entry_type = "Atualização"
            self.tl_custo_valor = ""
            self.tl_custo_categoria = "Operacional"
            self.tl_anexo_url = ""
            self.tl_anexo_nome = ""
            self.tl_submitting = False

        yield HubState.load_timeline(contrato)

    @rx.event(background=True)
    async def delete_timeline_entry(self, entry_id: str):
        contrato = ""
        async with self:
            contrato = self.timeline_entries[0]["contrato"] if self.timeline_entries else ""
        try:
            sb_delete("hub_timeline", filters={"id": entry_id})
        except Exception as e:
            logger.error(f"delete_timeline_entry error: {e}")
        if contrato:
            yield HubState.load_timeline(contrato)

    # ══════════════════════════════════════════════════════════════════════════
    # AGENTE DE ATIVIDADES — Insights pós-RDO
    # ══════════════════════════════════════════════════════════════════════════

    @rx.event(background=True)
    async def run_agente_atividades(self, contrato: str = "", rdo_id: str = "", force: bool = False):
        """
        Generates AI-powered activity insights based on schedule data + last RDO.
        Trigger: RDO submission (rdo_id muda) ou force=True (botão manual).
        Cache: não re-executa se insights existem para o mesmo contrato+rdo_id.
        """
        import json
        import asyncio
        from datetime import date, datetime
        from bomtempo.core.ai_client import ai_client
        from bomtempo.core.supabase_client import sb_select
        from bomtempo.core.circuit_breaker import ia_breaker

        # Guard: skip if already loading or cache hit (mesmo contrato + mesmo rdo_id)
        async with self:
            _already_loading = self.agente_loading
            _existing_contrato = self.agente_contrato
            _existing_rdo_id = self.agente_last_rdo_id
            _has_insights = len(self.agente_insights) > 0
            _target = contrato or self.agente_contrato or self.cron_contrato

        if _already_loading:
            return
        # Cache hit: mesmo contrato, mesmo rdo_id, tem insights — não gastar token
        if (not force and _has_insights and _target == _existing_contrato
                and rdo_id and rdo_id == _existing_rdo_id):
            return
        # Sem rdo_id e sem force: cache por contrato (comportamento anterior)
        if not force and not rdo_id and _has_insights and _target == _existing_contrato:
            return

        async with self:
            self.agente_loading = True
            self.agente_error = ""
            if contrato:
                self.agente_contrato = contrato
            _contrato = self.agente_contrato or self.cron_contrato

        if not _contrato:
            async with self:
                self.agente_loading = False
                self.agente_error = "Contrato não identificado."
            return

        loop = asyncio.get_running_loop()

        # Ensure cronograma is loaded (and fresh) for this contract
        # When triggered by a new RDO submission (rdo_id provided), always reload to pick up updated exec_qty
        async with self:
            _cron_rows = list(self.cron_rows)
            _cron_contrato = self.cron_contrato

        needs_reload = not _cron_rows or _cron_contrato != _contrato or bool(rdo_id)
        if needs_reload:
            yield HubState.load_cronograma(_contrato)
            await asyncio.sleep(2.0)
            async with self:
                _cron_rows = list(self.cron_rows)

        # Fetch last RDO header (for crew/climate context)
        def _fetch_last_rdo():
            try:
                rows = sb_select(
                    "rdo_master",
                    filters={"contrato": _contrato},
                    limit=1,
                    order="data.desc",
                ) or []
                return rows[0] if rows else {}
            except Exception:
                return {}

        # Fetch production history from hub_atividade_historico + rdo_atividades
        # This gives us quantidade/efetivo/dia per activity — the real productivity triangle
        def _fetch_production_data():
            try:
                from bomtempo.core.supabase_client import sb_rpc
                # Sanitize contrato code before SQL interpolation — strip quotes/semicolons
                # Contract codes are alphanumeric + hyphens (e.g. "CONT-001-2026"), never user-typed
                # but we sanitize defensively to prevent injection if a code is ever edited in DB.
                _safe_contrato = str(_contrato).replace("'", "''").replace(";", "")[:100]
                # Query: join rdo_atividades with rdo_master to get qty+efetivo+date per activity
                query = f"""
                    SELECT
                        ra.atividade,
                        ra.quantidade,
                        ra.unidade,
                        ra.efetivo,
                        rm.data,
                        rm.equipe_alocada,
                        rm.condicao_climatica
                    FROM rdo_atividades ra
                    JOIN rdo_master rm ON rm.id = ra.rdo_id
                    WHERE rm.contrato = '{_safe_contrato}'
                      AND ra.quantidade IS NOT NULL
                      AND ra.efetivo IS NOT NULL
                      AND ra.efetivo > 0
                    ORDER BY rm.data DESC
                    LIMIT 60
                """
                result = sb_rpc("execute_safe_query", {"query_string": query.strip()})
                return result or []
            except Exception:
                return []

        def _fetch_hist_data():
            try:
                rows = sb_select(
                    "hub_atividade_historico",
                    filters={"contrato": _contrato},
                    limit=50,
                    order="created_at.desc",
                ) or []
                return rows
            except Exception:
                return []

        try:
            last_rdo, prod_rows, hist_rows = await asyncio.gather(
                loop.run_in_executor(get_db_executor(), _fetch_last_rdo),
                loop.run_in_executor(get_db_executor(), _fetch_production_data),
                loop.run_in_executor(get_db_executor(), _fetch_hist_data),
            )
        except Exception as _gather_err:
            logger.error(f"generate_agente_insights: db gather falhou: {_gather_err}")
            async with self:
                self.agente_loading = False
                self.agente_error = "Erro ao buscar dados. Tente novamente."
            return

        try:
            # Build context for AI
            today_dt = date.today()
            today = today_dt.isoformat()

            # Activity summary (top 20 most relevant rows)
            micros = [r for r in _cron_rows if r.get("nivel", "") == "micro"]
            if not micros:
                micros = _cron_rows  # fallback: all rows

            def _ativ_summary(rows):
                lines = []
                for r in rows[:25]:
                    pct = r.get("conclusao_pct", "0")
                    name = r.get("atividade", "")
                    termino = r.get("termino_previsto", "")
                    inicio = r.get("inicio_previsto", "")
                    critico = r.get("critico", "0")
                    total_qty = r.get("total_qty", "")
                    exec_qty = r.get("exec_qty", "")
                    unidade = r.get("unidade", "")
                    responsavel = r.get("responsavel", "")
                    status = r.get("status_atividade", "nao_iniciada")
                    peso = r.get("peso_pct", "")
                    prod_info = f", qty: {exec_qty}/{total_qty} {unidade}".strip() if total_qty else ""
                    line = (
                        f"- [{r.get('fase','')}.{r.get('fase_macro','')}] {name}"
                        f" | {pct}% | {inicio}→{termino}"
                        f" | resp: {responsavel} | status: {status}"
                        f" | critico: {critico} | peso: {peso}%{prod_info}"
                    )
                    lines.append(line)
                return "\n".join(lines)

            ativ_text = _ativ_summary(micros)

            # Last RDO summary — from rdo_master (correct table)
            rdo_text = ""
            if last_rdo:
                obs = (last_rdo.get('observacoes') or '').strip()
                orient = (last_rdo.get('orientacao') or '').strip()
                rdo_text = (
                    f"Data de referência: {last_rdo.get('data','')}\n"
                    f"Equipe alocada: {last_rdo.get('equipe_alocada','')}\n"
                    f"Condição climática: {last_rdo.get('condicao_climatica','')}\n"
                    f"Houve chuva: {last_rdo.get('houve_chuva','')}\n"
                    f"Houve interrupção: {last_rdo.get('houve_interrupcao','')}\n"
                    f"Motivo interrupção: {last_rdo.get('motivo_interrupcao','')}\n"
                    f"Observações: {obs or 'Nenhuma'}\n"
                    + (f"⚠️ ORIENTAÇÕES/PENDÊNCIAS DO MESTRE: {orient}\n" if orient else "")
                    + f"Houve acidente: {last_rdo.get('houve_acidente','')}\n"
                )

            # ── Compute derived metrics for richer prompt context ─────────────────
            from datetime import date as _date_cls
            late_lines: list = []
            critical_late: list = []
            ahead_lines: list = []
            not_started_critical: list = []
            total_peso = 0.0
            realizado_peso = 0.0

            for r in micros:
                termino_iso = (r.get("data_fim_real_iso") or r.get("termino_previsto", ""))[:10]
                pct_str = r.get("conclusao_pct", "0") or "0"
                critico = r.get("critico", "0")
                name = r.get("atividade", "")
                status = r.get("status_atividade", "nao_iniciada")
                try:
                    pct_val = float(pct_str)
                except (ValueError, TypeError):
                    pct_val = 0.0
                try:
                    peso_val = float(r.get("peso_pct", "0") or "0")
                    total_peso += peso_val
                    realizado_peso += peso_val * pct_val / 100.0
                except (ValueError, TypeError):
                    pass
                if pct_val < 100.0 and termino_iso:
                    try:
                        term_dt = _date_cls.fromisoformat(termino_iso)
                        delta = (today_dt - term_dt).days  # type: ignore[attr-defined]
                        if delta > 0:
                            tag = " [CRÍTICO]" if critico == "1" else ""
                            late_lines.append(f"  - {name}: {delta}d atrasada{tag}")
                            if critico == "1":
                                critical_late.append((name, delta))
                        elif delta < -30 and pct_val > 10:
                            ahead_lines.append(f"  - {name}: {abs(delta)}d de folga, {pct_val:.0f}% concluído")
                    except ValueError:
                        pass
                if critico == "1" and status == "nao_iniciada":
                    inicio_iso = r.get("inicio_previsto", "")[:10]
                    try:
                        ini_dt = _date_cls.fromisoformat(inicio_iso)
                        if ini_dt <= today_dt:  # type: ignore[attr-defined]
                            not_started_critical.append(name)
                    except ValueError:
                        pass

            spi_geral = round(realizado_peso / total_peso, 2) if total_peso > 0 else None
            # Calcular SPI corrigido: apenas sobre atividades que já deveriam ter iniciado
            # Atividades futuras (inicio > hoje) ainda NÃO deveriam estar em andamento — excluir do SPI
            active_peso = 0.0
            active_realizado = 0.0
            for r in micros:
                ini_iso = r.get("inicio_previsto", "")[:10] or r.get("inicio_iso", "")[:10]
                if not ini_iso:
                    continue
                try:
                    ini_dt = _date_cls.fromisoformat(ini_iso)
                except ValueError:
                    continue
                if ini_dt > today_dt:
                    continue  # atividade futura — não penaliza SPI
                try:
                    peso_val = float(r.get("peso_pct", "0") or "0")
                    pct_val = float(r.get("conclusao_pct", "0") or "0")
                    active_peso += peso_val
                    active_realizado += peso_val * pct_val / 100.0
                except (ValueError, TypeError):
                    pass
            if active_peso > 0:
                spi_ativo = round(active_realizado / active_peso, 2)
                spi_line = (
                    f"SPI das atividades ativas (já iniciadas): {spi_ativo} "
                    f"(IMPORTANTE: {len([r for r in micros if (r.get('inicio_previsto','')[:10] or r.get('inicio_iso','')[:10]) > today_dt.isoformat()])} "
                    f"atividades são futuras e NÃO devem ser penalizadas no SPI)"
                )
            elif spi_geral is not None:
                spi_line = f"SPI global do projeto: {spi_geral} (inclui atividades ainda não iniciadas)"
            else:
                spi_line = "SPI: dados insuficientes"

            # Calcular dias_sem_rdo para diferenciar "atrasado" de "aguardando lançamento"
            dias_sem_rdo = None
            rdo_status_line = ""
            if last_rdo:
                try:
                    _last_rdo_date_str = (last_rdo.get("data") or "")[:10]
                    if _last_rdo_date_str:
                        _last_rdo_dt = _date_cls.fromisoformat(_last_rdo_date_str)
                        dias_sem_rdo = (today_dt - _last_rdo_dt).days
                        if dias_sem_rdo == 0:
                            rdo_status_line = "⏳ RDO de HOJE já recebido — dados atualizados."
                        elif dias_sem_rdo == 1:
                            rdo_status_line = f"⏳ Aguardando RDO de hoje ({today_str}). Último: {_last_rdo_date_str}. SPI pode estar desatualizado por 1 dia — NÃO é atraso."
                        elif dias_sem_rdo <= 3:
                            rdo_status_line = f"⚠️ {dias_sem_rdo} dias sem RDO (último: {_last_rdo_date_str}). Dados de progresso podem estar desatualizados."
                        else:
                            rdo_status_line = f"🔴 {dias_sem_rdo} dias sem RDO (último: {_last_rdo_date_str}). Comunicação com campo comprometida."
                except (ValueError, AttributeError):
                    pass
            late_summary = "\n".join(late_lines[:12]) if late_lines else "  Nenhuma atividade com prazo vencido detectada."
            ahead_summary = "\n".join(ahead_lines[:5]) if ahead_lines else "  Nenhuma."
            not_started_summary = "\n".join(f"  - {n}" for n in not_started_critical[:5]) if not_started_critical else "  Nenhuma."

            # ── Crew context from rdo_master (correct fields) ─────────────────────
            crew_context = ""
            if last_rdo:
                try:
                    alocada = int(last_rdo.get("equipe_alocada") or 0)
                    clima = last_rdo.get("condicao_climatica", "")
                    chuva = last_rdo.get("houve_chuva", False)
                    interrupcao = last_rdo.get("houve_interrupcao", False)
                    motivo = last_rdo.get("motivo_interrupcao", "")
                    acidente = last_rdo.get("houve_acidente", False)
                    rdo_date = last_rdo.get("data", "")

                    parts = [f"Último RDO ({rdo_date}): {alocada} pessoas alocadas, clima: {clima}"]
                    if chuva:
                        parts.append("⚠️ Houve chuva")
                    if interrupcao:
                        parts.append(f"⚠️ Houve interrupção: {motivo or 'motivo não informado'}")
                    if acidente:
                        parts.append("🔴 ACIDENTE REGISTRADO no último RDO")
                    crew_context = " | ".join(parts)
                except (ValueError, TypeError):
                    crew_context = f"Equipe do último RDO: {last_rdo.get('equipe_alocada','?')} pessoas."

            # ── Real productivity from rdo_atividades (quantidade + efetivo + data) ─
            # Fonte: hub_atividade_historico (producao_dia, exec_qty, total_qty, unidade)
            # + rdo_atividades (quantidade, unidade, efetivo por atividade por dia)
            productivity_context = ""
            try:
                prod_lines = []
                # Group by atividade → compute taxa = quantidade / efetivo per day
                from collections import defaultdict
                activity_rates: dict = defaultdict(list)  # atividade → [(qty, efetivo, data)]

                for row in prod_rows:
                    act_name = (row.get("atividade") or "").strip()
                    qty = row.get("quantidade")
                    efetivo = row.get("efetivo")
                    data = row.get("data", "")
                    unidade = row.get("unidade", "")
                    if not act_name or qty is None or efetivo is None:
                        continue
                    try:
                        qty_f = float(qty)
                        efetivo_i = int(efetivo)
                        if efetivo_i > 0 and qty_f > 0:
                            activity_rates[act_name].append({
                                "qty": qty_f, "efetivo": efetivo_i,
                                "taxa": round(qty_f / efetivo_i, 2),
                                "data": data, "unidade": unidade
                            })
                    except (ValueError, TypeError):
                        pass

                for act_name, samples in activity_rates.items():
                    if not samples:
                        continue
                    unidade = samples[0]["unidade"]
                    avg_taxa = round(sum(s["taxa"] for s in samples) / len(samples), 2)
                    avg_efetivo = round(sum(s["efetivo"] for s in samples) / len(samples), 1)
                    avg_qty = round(sum(s["qty"] for s in samples) / len(samples), 1)
                    n = len(samples)

                    # Saldo restante da atividade (da tabela hub_atividades via cron_rows)
                    cron_act = next((x for x in micros if x.get("atividade", "").strip() == act_name), None)
                    saldo_str = ""
                    recovery_str = ""
                    if cron_act:
                        try:
                            total_q = float(cron_act.get("total_qty") or 0)
                            exec_q = float(cron_act.get("exec_qty") or 0)
                            remaining = total_q - exec_q
                            if remaining > 0 and avg_taxa > 0:
                                # Opção A: recuperar com equipe histórica média
                                dias_com_media = round(remaining / (avg_taxa * avg_efetivo), 1)
                                # Opção B: recuperar em prazo reduzido (2/3 dos dias) → workers necessários
                                prazo_target = max(1, dias_com_media * 0.67)
                                workers_needed = round(remaining / (avg_taxa * prazo_target), 1)
                                saldo_str = f" | saldo: {remaining:.1f} {unidade}"
                                recovery_str = (
                                    f" | RECOVERY: A) {dias_com_media}d com {avg_efetivo:.0f}p "
                                    f"B) {round(prazo_target)}d com {workers_needed:.0f}p"
                                )
                        except (ValueError, TypeError):
                            pass

                    # Efetivo alocado planejado vs real para granularidade
                    efetivo_plan_str = ""
                    if cron_act:
                        try:
                            ef_plan = int(cron_act.get("efetivo_alocado", 0) or 0)
                            if ef_plan > 0:
                                diff = avg_efetivo - ef_plan
                                status = "OK" if abs(diff) <= 1 else (f"+{diff:.0f}p extra" if diff > 0 else f"{abs(diff):.0f}p faltando")
                                efetivo_plan_str = f" [planejado: {ef_plan}p → real: {avg_efetivo:.0f}p → {status}]"
                        except (ValueError, TypeError):
                            pass

                    prod_lines.append(
                        f"  - [{act_name}] taxa histórica: {avg_taxa} {unidade}/pessoa/dia "
                        f"(média {n} RDOs, {avg_efetivo:.0f}p, {avg_qty:.1f} {unidade}/dia)"
                        f"{efetivo_plan_str}{saldo_str}{recovery_str}"
                    )

                # Also enrich with hub_atividade_historico producao_dia for any gaps
                hist_acts_seen = {s.split("] ")[0].lstrip("  - [") for s in prod_lines}
                hist_by_act: dict = defaultdict(list)
                for h in hist_rows:
                    act_id = h.get("atividade_id", "")
                    prod_dia = h.get("producao_dia")
                    unidade_h = h.get("unidade", "")
                    if prod_dia and act_id:
                        hist_by_act[act_id].append({"prod": float(prod_dia), "unidade": unidade_h})

                if prod_lines:
                    productivity_context = "\n".join(prod_lines)
                else:
                    productivity_context = "  Sem dados de quantidade+efetivo nos RDOs para calcular taxa."
            except Exception as _e:
                logger.warning(f"Agente productivity calc error: {_e}")
                productivity_context = "  Erro ao calcular dados de produtividade."

            today_str: str = today  # already defined as date.today().isoformat()

            prompt = f"""Você é o Agente de Atividades da plataforma Bomtempo — especialista sênior em engenharia civil, gestão de obras e análise de cronograma. Sua audiência é o gestor de obra: ele lê seu insight e em 30 segundos sabe exatamente o que fazer.

MISSÃO: gerar entre 4 e 6 insights CIRÚRGICOS e ACIONÁVEIS. O gestor deve ler e já ter a decisão na mão.

═══ REGRAS OBRIGATÓRIAS (violá-las invalida o insight) ═══
1. Cada insight DEVE citar ao menos 1 número real (dias, %, unidades, trabalhadores, R$)
2. Para atrasos com dados de produtividade disponíveis: CALCULE a equipe necessária para recuperação.
   Fórmula de workforce recovery: meta_recuperação ÷ taxa_produção_pessoa_dia = workers_necessários
   Apresente 2 opções: (A) recuperar em N dias com X pessoas, (B) alongar prazo e novo término
3. Para atividades ADIANTADAS: calcule o ganho real e sugira realocação com nome e quantidade
4. Atividades com critico=1 atrasadas → type="delay" ou "alert" com priority="high" OBRIGATORIAMENTE
5. Não invente dados. Se estimar, sinalize com "(estimado)". Se os dados são insuficientes para o cálculo, diga qual informação falta
6. O campo "atividade" deve ser o nome EXATO da atividade do cronograma (copie literalmente), ou '' se o insight for geral
7. Insights POSITIVOS são tão importantes quanto negativos — reconheça e capitalize sobre o que está indo bem
8. CRÍTICO — NÃO confunda ausência de dados com atraso: atividades com início_previsto > hoje são FUTURAS e não estão atrasadas. Só considere atrasada uma atividade cujo termino_previsto < hoje E pct < 100. Atividades futuras sem exec_qty são NORMAIS — obra não começou essa fase ainda.
9. SPI do contexto abaixo é calculado APENAS sobre atividades já iniciadas (início ≤ hoje). Não extrapole para "equipe insuficiente" se as atividades ativas estão todas concluídas ou no ritmo.
10. CRÍTICO — "Dias sem RDO" no contexto NÃO é automaticamente atraso: 0 dias = dados atuais; 1 dia = aguardando RDO de hoje (normal); ≥2 dias = possível atraso de comunicação. Só gere alerta de atraso operacional se termino_previsto < hoje E dias_sem_rdo >= 2.

═══ CONTEXTO DO PROJETO ═══
Contrato: {_contrato}
Data de hoje: {today_str}
{spi_line}
{rdo_status_line}

═══ ATIVIDADES COM PRAZO VENCIDO ═══
{late_summary}

═══ ATIVIDADES ADIANTADAS (oportunidade de realocação) ═══
{ahead_summary}

═══ ATIVIDADES CRÍTICAS NÃO INICIADAS (deveriam ter começado) ═══
{not_started_summary}

═══ SITUAÇÃO DE EQUIPE (último RDO) ═══
{crew_context or "Sem dados de equipe no RDO."}

═══ PRODUTIVIDADE HISTÓRICA POR ATIVIDADE (fonte: rdo_atividades + hub_atividade_historico — use para calcular equipe necessária) ═══
{productivity_context or "  Sem histórico de produtividade disponível."}
INSTRUÇÃO: use os dados de produtividade acima para calcular precisamente quantos trabalhadores e por quantos dias são necessários para recuperar atrasos. Formato esperado: "Meta: X unid em Y dias úteis → Z unid/dia → W pessoas (taxa histórica: V unid/pessoa/dia)"

═══ CRONOGRAMA COMPLETO (formato: [fase.subfase] nome | %concluído | início→término | responsável | status | crítico | peso% | qty exec/total unidade) ═══
{ativ_text or "Nenhuma atividade cadastrada."}

═══ ÚLTIMO RDO REGISTRADO ═══
{rdo_text or "Nenhum RDO encontrado para este contrato."}

Responda APENAS com um array JSON válido (sem markdown, sem ```json, sem explicações fora do array):
[
  {{
    "type": "delay|ahead|crew|weather|optimize|alert",
    "priority": "high|medium|low",
    "icon": "nome-do-icone-lucide",
    "title": "Insight com número real (max 60 chars)",
    "atividade": "nome exato da atividade ou ''",
    "body": "2-3 frases com números específicos + recomendação acionável concreta"
  }}
]

Exemplos de body de ALTA QUALIDADE:
EXEMPLOS DE BODY — ALTA QUALIDADE (padrão enterprise):
- "Fundações Bloco A: 18d atrasada, saldo de 240m² (60% do total). Taxa histórica: 5 pessoas → 20m²/dia. Opção A: manter prazo — precisa de 9 pessoas por 3 semanas (240m² ÷ 20m²/dia = 12 dias, 9p = 26m²/dia cobrindo folga). Opção B: realocar equipe atual e estender término em 18 dias úteis."
- "Estrutura Metálica: 78% concluído vs meta 65% — 13pp adiantada, 12 dias de folga. Realocar 2 dos 8 soldadores para Instalações Elétricas (42% vs meta 55%) normaliza portfólio sem comprometer prazo."
- "Equipe 40% abaixo do planejado nos últimos 3 RDOs (6/10 presentes em média). Perda acumulada estimada: 4,2pp de avanço/semana. Se persistir, 'Concretagem Pilares' entra em zona crítica em 9 dias. Ação: contratar 2 serventes temporários ou acionar banco de horas."
- "POSITIVO: 'Instalação Hidráulica' concluiu 95% com 3 dias de antecedência. Equipe de 4 encanadores está disponível a partir de amanhã — alocar imediatamente para 'Instalação Elétrica Térreo' (apenas 30% concluída, prazo em 15 dias)."

Tipos:
- delay: atividade atrasada ou em risco iminente de atraso (SEMPRE com cálculo de recovery se possível)
- ahead: atividade adiantada — oportunidade de realocar recursos (SEMPRE com quem realocar e para onde)
- crew: dimensionamento de equipe — gap ou excesso (SEMPRE com número exato de pessoas faltantes/extras)
- weather: impacto climático registrado no RDO ou risco meteorológico
- optimize: oportunidade de paralelismo, sequenciamento ou compressão de prazo
- alert: alerta crítico geral sem atividade específica

Ícones Lucide: alert-triangle, trending-up, trending-down, users, cloud-rain, calendar-check, zap, clock, hard-hat, wrench, check-circle, flame, shield-alert, target, layers"""

            def _call_ai():
                try:
                    return ai_client.query(
                        [{"role": "user", "content": prompt}],
                        model="gpt-4o",
                        username="agente_atividades",
                    )
                except Exception as e:
                    logger.error(f"Agente IA error: {e}")
                    return ""

            raw = ""
            if not ia_breaker.is_open():
                try:
                    raw = await asyncio.wait_for(
                        loop.run_in_executor(get_ai_executor(), _call_ai),
                        timeout=40.0,
                    )
                    ia_breaker.record_success()
                except asyncio.TimeoutError:
                    ia_breaker.record_failure()
                    raw = ""
                except Exception as e:
                    ia_breaker.record_failure(e)
                    raw = ""

            # Parse JSON from response
            insights: list = []
            if raw:
                try:
                    # Strip possible markdown fences
                    text = raw.strip()
                    if text.startswith("```"):
                        text = text.split("```")[1]
                        if text.startswith("json"):
                            text = text[4:]
                    parsed = json.loads(text.strip())
                    if isinstance(parsed, list):
                        insights = [
                            {
                                "type": str(item.get("type", "alert")),
                                "priority": str(item.get("priority", "medium")),
                                "icon": str(item.get("icon", "zap")),
                                "title": str(item.get("title", ""))[:80],
                                "atividade": str(item.get("atividade", "")),
                                "body": str(item.get("body", "")),
                            }
                            for item in parsed
                            if isinstance(item, dict)
                        ]
                except Exception as parse_err:
                    logger.error(f"Agente parse error: {parse_err}\nraw={raw[:200]}")

            now_brt = datetime.now(_BRT).strftime("%d/%m/%Y %H:%M")

            async with self:
                self.agente_loading = False
                self.agente_last_updated = now_brt
                self.agente_contrato = _contrato
                if rdo_id:
                    self.agente_last_rdo_id = rdo_id
                if insights:
                    self.agente_insights = insights
                    self.agente_error = ""
                else:
                    self.agente_error = "Não foi possível gerar insights no momento. Tente novamente."

            # Persist new insights to Supabase so they survive page navigation
            if insights:
                _insights_snapshot = list(insights)
                _contrato_snap = _contrato
                _rdo_snap = rdo_id or ""
                _ts_snap = now_brt
                def _save():
                    try:
                        from bomtempo.core.supabase_client import sb_upsert
                        import json as _json
                        # insights column is jsonb — pass as JSON string, PostgREST will cast it
                        sb_upsert(
                            "agente_insights",
                            {
                                "contrato":    _contrato_snap,
                                "insights":    _json.dumps(_insights_snapshot, ensure_ascii=False),
                                "last_rdo_id": _rdo_snap,
                                "updated_at":  _ts_snap,
                            },
                            on_conflict="contrato",
                        )
                        logger.info(f"agente_insights saved for {_contrato_snap}")
                    except Exception as _e:
                        logger.warning(f"agente_insights save failed: {_e}")
                import threading as _threading
                _threading.Thread(target=_save, daemon=True).start()

        except Exception as _agente_err:
            logger.error(f"run_agente_atividades error: {_agente_err}", exc_info=True)
        finally:
            # Garante que agente_loading SEMPRE volta pra False — mesmo em crash/CancelledError
            async with self:
                if self.agente_loading:
                    self.agente_loading = False

    @rx.event(background=True)
    async def load_persisted_insights(self, contrato: str):
        """
        Load persisted insights from Supabase for the given contract.
        Called when switching projects — shows saved insights instantly
        without triggering a new AI call.
        Sets agente_loading=True while fetching so stale data from the previous
        project is never exposed.
        """
        import json as _json
        import asyncio

        # Clear stale data immediately and show loading spinner
        async with self:
            self.agente_insights = []
            self.agente_error = ""
            self.agente_contrato = contrato
            self.agente_last_rdo_id = ""
            self.agente_last_updated = ""
            self.agente_loading = True

        loop = asyncio.get_running_loop()

        def _fetch():
            try:
                from bomtempo.core.supabase_client import sb_select
                rows = sb_select(
                    "agente_insights",
                    filters={"contrato": contrato},
                    limit=1,
                )
                return rows[0] if rows else None
            except Exception as _e:
                logger.warning(f"agente_insights load failed: {_e}")
                return None

        try:
            row = await loop.run_in_executor(get_db_executor(), _fetch)
        except Exception as e:
            logger.error(f"load_persisted_insights executor error: {e}", exc_info=True)
            async with self:
                self.agente_loading = False
            return

        if not row:
            # No saved insights for this contract — leave empty so user can generate
            async with self:
                self.agente_loading = False
            return

        async with self:
            try:
                raw_insights = row.get("insights", "[]")
                # Handle both jsonb (already a list) and text (needs parsing)
                if isinstance(raw_insights, list):
                    saved = raw_insights
                else:
                    saved = _json.loads(raw_insights)
                if isinstance(saved, list) and saved:
                    self.agente_insights = [
                        {
                            "type":      str(x.get("type", "alert")),
                            "priority":  str(x.get("priority", "medium")),
                            "icon":      str(x.get("icon", "zap")),
                            "title":     str(x.get("title", ""))[:80],
                            "atividade": str(x.get("atividade", "")),
                            "body":      str(x.get("body", "")),
                        }
                        for x in saved if isinstance(x, dict)
                    ]
                    self.agente_last_rdo_id = row.get("last_rdo_id", "")
                    self.agente_last_updated = row.get("updated_at", "")
            except Exception as _e:
                logger.warning(f"agente_insights parse failed: {_e}")
            finally:
                self.agente_loading = False

    def reload_after_rdo(self, contrato: str = ""):
        """Invalida o cache do cronograma após submit de um RDO.
        Chamado via yield em rdo_state.execute_submit para garantir que o hub
        mostre dados atualizados sem que o usuário precise sair e voltar.
        Se o cronograma desse contrato já estava carregado, dispara reload."""
        self._cron_forecast_cache = []
        if contrato and self.cron_contrato == contrato:
            # Cronograma deste contrato estava visível — recarrega silenciosamente
            return HubState.load_cronograma(contrato)

    def force_run_agente(self, contrato: str = ""):
        """Clear existing insights and force a fresh Agente run (botão manual)."""
        self.agente_insights = []
        self.agente_error = ""
        self.agente_last_rdo_id = ""  # invalida cache
        _c = contrato or self.agente_contrato or self.cron_contrato
        return HubState.run_agente_atividades(_c, force=True)

    def reset_for_logout(self):
        """Limpa todo o estado sensível do Hub ao fazer logout.
        Evita flicker de dados de outro usuário/tenant ao fazer login na sequência.
        Chamado via yield em GlobalState.logout."""
        self.cron_rows = []
        self._cron_forecast_cache = []
        self.cron_contrato = ""
        self.cron_fase_filter = ""
        self.cron_search = ""
        self.cron_search_input = ""
        self._audit_loaded_contrato = ""
        self._timeline_loaded_contrato = ""
        self.audit_images = []
        self.timeline_entries = []
        self.agente_insights = []
        self.agente_loading = False
        self.agente_error = ""
        self.agente_contrato = ""
        self.agente_last_rdo_id = ""
        self.agente_last_updated = ""
